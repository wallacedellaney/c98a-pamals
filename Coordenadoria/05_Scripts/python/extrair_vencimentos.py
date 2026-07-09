"""Lê a aba C-98U8 da planilha "Vencimentos" (TMOT — controle de vencimento de
itens por hora, pouso ou calendário) e gera 02_Dados_Tratados/base_vencimentos_tratada.xlsx.

Fonte: 01_Bases_Originais/Vencimentos/Vencimentos_C-98U8.xlsx (cópia local só
com as colunas usadas — ver 00_Instrucoes/vencimentos.md).

A coluna DISPONIBILIDADE vem em formatos diferentes na planilha original,
dependendo do tipo de vencimento do item (não tem uma coluna própria dizendo
o tipo — é preciso inferir pelo formato do valor):
* timedelta (Excel formata como [h]:mm:ss) -> vencimento por HORA. Convertido
  para um número de horas (positivo ou negativo).
* número puro (int/float) -> vencimento por POUSO (quantidade de pousos).
* texto em meses ("-368m"), dias ("-24d", "23d") ou mês+dia combinados
  ("1me8d") -> vencimento por CALENDÁRIO. Convertido para um total de DIAS
  (mês = 30 dias, aproximado) pra ter uma unidade só de comparação/filtro;
  o texto original (ex.: "1me8d") é mantido à parte pra exibição.

Em qualquer um dos casos, DISPONIBILIDADE negativa = item VENCIDO (já
passou do limite) por aquela quantidade de horas/pousos/dias.
"""

import re
from datetime import datetime

import openpyxl
import pandas as pd

from common import BASES_ORIGINAIS, DADOS_TRATADOS, ESTADO_ATUALIZACOES, registrar_log
from shared import drive_sync, estado

FONTE = BASES_ORIGINAIS / "Vencimentos" / "Vencimentos_C-98U8.xlsx"

# A planilha real "Vencimentos" tem 10 abas (uma por tipo de aeronave — C-97A2,
# A-29T2, C-98U8, G-19U2, IU-93A6, T-25T9, T-27T1, PNs TGCC, Lixos) — a aba
# "ativa" por padrão é "C-97A2", não a nossa. Sempre referenciar pelo nome.
ABA = "C-98U8"

# Ver 00_Instrucoes/vencimentos.md — planilha Google nativa "Vencimentos",
# compartilhada com a conta de serviço em 2026-07-09.
DRIVE_FILE_ID = "178vQ-lRP52sw30kQArqcsQGXfj2OLblaFCgjIXWFIl8"

RE_MES_DIA = re.compile(r"^(-?)(\d+)me(\d+)d$", re.IGNORECASE)
RE_MES = re.compile(r"^(-?\d+(?:[.,]\d+)?)\s*m$", re.IGNORECASE)
RE_DIA = re.compile(r"^(-?\d+(?:[.,]\d+)?)\s*d$", re.IGNORECASE)


def _classificar_disponibilidade(valor):
    """Retorna (tipo, valor_numerico, texto_original) a partir do valor bruto
    da coluna DISPONIBILIDADE. Para "Calendário", valor_numerico é sempre em
    dias (mês aproximado como 30 dias)."""
    if valor is None:
        return None, None, None
    if hasattr(valor, "total_seconds"):  # timedelta -> horas
        return "Hora", round(valor.total_seconds() / 3600, 1), None
    if isinstance(valor, (int, float)):
        return "Pouso", valor, None
    if isinstance(valor, str):
        texto = valor.strip()
        m = RE_MES_DIA.match(texto)
        if m:
            sinal = -1 if m.group(1) == "-" else 1
            dias = sinal * (int(m.group(2)) * 30 + int(m.group(3)))
            return "Calendário", dias, texto
        m = RE_MES.match(texto)
        if m:
            return "Calendário", round(float(m.group(1).replace(",", ".")) * 30), texto
        m = RE_DIA.match(texto)
        if m:
            return "Calendário", round(float(m.group(1).replace(",", "."))), texto
    return "Desconhecido", None, str(valor)


def extrair():
    inconsistencias = []
    wb = openpyxl.load_workbook(FONTE, data_only=True)
    ws = wb[ABA]

    linhas = []
    for r in range(2, ws.max_row + 1):
        disponibilidade, data_venc, inspecao, pn, sn, matricula, operador, nomenclatura = (
            ws.cell(row=r, column=c).value for c in range(1, 9)
        )
        if matricula is None and pn is None:
            continue

        tipo, valor_numerico, texto_original = _classificar_disponibilidade(disponibilidade)
        if tipo == "Desconhecido":
            inconsistencias.append(
                f"Linha {r}: DISPONIBILIDADE em formato não reconhecido ({disponibilidade!r}) — item ignorado."
            )
            continue
        if tipo is None:
            inconsistencias.append(f"Linha {r}: DISPONIBILIDADE vazia — item ignorado.")
            continue

        linhas.append({
            "matricula": str(matricula) if matricula is not None else None,
            "operador": operador,
            "pn": str(pn) if pn is not None else None,
            "sn": str(sn) if sn is not None else None,
            "nomenclatura": nomenclatura,
            "inspecao": str(inspecao) if inspecao is not None else None,
            "tipo_vencimento": tipo,
            "disponibilidade_valor": valor_numerico,
            "disponibilidade_texto": texto_original,
            "data_vencimento": data_venc.date() if hasattr(data_venc, "date") else None,
            "vencido": valor_numerico < 0 if valor_numerico is not None else None,
        })

    df = pd.DataFrame(linhas)
    return df, inconsistencias


def main():
    DADOS_TRATADOS.mkdir(parents=True, exist_ok=True)
    df, inconsistencias = extrair()

    destino = DADOS_TRATADOS / "base_vencimentos_tratada.xlsx"
    with pd.ExcelWriter(destino) as writer:
        df.to_excel(writer, index=False, sheet_name="TMOT")

    registrar_log(
        nome_execucao="extrair_vencimentos",
        arquivos_lidos=[str(FONTE)],
        arquivos_gerados=[str(destino)],
        inconsistencias=inconsistencias,
    )

    print(f"{len(df)} itens de vencimento (TMOT) -> {destino}")
    if inconsistencias:
        print(f"{len(inconsistencias)} inconsistência(s) encontrada(s), ver log em 06_Logs/.")

    return df


def atualizar_do_drive():
    """Busca a versão mais recente direto do Google Drive, sobrescreve a
    cópia local e reprocessa. Ver 00_Instrucoes/atualizacoes.md."""
    try:
        metadados = drive_sync.obter_metadados(DRIVE_FILE_ID)
        conteudo = drive_sync.baixar_arquivo(DRIVE_FILE_ID, exportar_como=drive_sync.XLSX_MIME)
        FONTE.parent.mkdir(parents=True, exist_ok=True)
        FONTE.write_bytes(conteudo)
        df = main()
        estado.atualizar_estado(
            ESTADO_ATUALIZACOES, "vencimentos_tmot",
            remote_modified_time=metadados["modifiedTime"],
            local_updated_at=datetime.now().isoformat(),
            status="atualizado",
            record_count=len(df),
            last_error=None,
        )
    except Exception as e:
        estado.atualizar_estado(ESTADO_ATUALIZACOES, "vencimentos_tmot", status="erro", last_error=str(e))
        raise
    return estado.obter_entrada(ESTADO_ATUALIZACOES, "vencimentos_tmot")


if __name__ == "__main__":
    import sys
    if "--atualizar-do-drive" in sys.argv:
        atualizar_do_drive()
    else:
        main()
