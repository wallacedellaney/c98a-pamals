"""Compila os "Controle de Vencimentos" de cada operador/base (CLA, DACTA II,
PAMA-LS, BACO, BANT, BACG, BABR, BAMN, BABE — ver 00_Instrucoes/vencimentos.md)
num único arquivo tratado, junto com o mês/arquivo de origem de cada linha.

Cada operador manda um arquivo com nome e formato próprio (alguns .xlsx,
alguns .ods, alguns tudo num arquivo só). Por isso o registro abaixo é
explícito por operador — sempre que o Claude buscar um arquivo novo no Drive
(o mais recente; se um operador não mandou nesse mês, usa o mês anterior dele
até achar), atualiza esse registro e salva o arquivo em
01_Bases_Originais/Vencimentos/Operadores/<OPERADOR>/.

Formatos incorporados até agora:
* CLA, DACTA II, BABR, BABE, BACO — planilha limpa com cabeçalho
  ESPECIALIDADE/NOMENCLATURA DO ITEM/PN/SN/AERONAVE/DISPONIBILIDADE/DATA
  ESTIMADA DO VENCIMENTO e seções "POR HORA/POR POUSO/POR CALENDÁRIO" marcadas
  (embora nem sempre à risca — por isso o tipo real de cada linha é decidido
  pelo FORMATO do valor de DISPONIBILIDADE, não pela seção onde está).
* PAMA-LS — sem seções, matrícula fixa no cabeçalho (só a aeronave 2704).
* BAMN — vem numa aba "VENCIMENTO" dentro do arquivo "Diagonal de Manutenção"
  (não um arquivo próprio). Ordem de colunas diferente (NOMENCLATURA/PN/SN/
  ESPECIALIDADE/AERONAVE/...), usa células **mescladas verticalmente** pra
  nomenclatura (o valor vale pra várias linhas de baixo — é preciso "puxar pra
  baixo" o último valor real da coluna), matrícula sem prefixo "FAB" (só o
  número), e tem formatos de valor extras: sufixo "A"/"Anos" (anos), "Nm e Nd"
  (mês + dia com espaços), datas "MM/AA" (só mês/ano, sem dia).
* BACG — a fonte de julho/2026 veio íntegra em .xlsx, com a aba "Controle de
  Vencimentos" e AERONAVE sem prefixo "FAB" (igual à BAMN). Substituiu o CSV
  reconstruído de abril. Tem itens com DISPONIBILIDADE "O/C" (On Condition —
  sem vencimento programado), tratados como tipo próprio ("Condição"), não
  como inconsistência.

* BANT — arquivo único (Diagonal+MAPEM+Vencimento) com 3 abas; a aba
  "Controle de Vencimento de Itens" já usa a ordem de colunas padrão, só que
  com AERONAVE em número puro (sem "FAB"). Trouxe formatos de valor novos:
  hora com ponto de milhar ("4.938:00"), hora anotada com equivalente em
  meses entre parênteses ("1.337:20 (19,8 M)" — o parêntese é só informativo
  e é descartado), "ANV S/ MOTOR"/variações de "NÃO INSTALADO" (item não
  instalado, sem vencimento aplicável) e "VENCIDA" (sabidamente vencido, mas
  sem valor informado — `tipo_vencimento="Vencido (sem valor)"`, vencido=True
  mesmo sem número). A coluna de data também trouxe "JUL / 27" (mês/ano com
  espaço) e 4 datas claramente erradas (anos 1930/1931/1935 — bug de
  formatação na origem), tratadas como data ausente (não inventamos).

Todos os 9 operadores confirmados (BAMN, BABE, CLA, BANT, BABR, PAMA-LS,
DACTA II, BACO, BACG) estão incorporados.
"""

import re

import openpyxl
import pandas as pd

from common import BASES_ORIGINAIS, DADOS_TRATADOS, registrar_log
from vencimentos_parse import classificar_disponibilidade, parse_data_vencimento, vencido_de

OPERADORES_DIR = BASES_ORIGINAIS / "Vencimentos" / "Operadores"

RE_AERONAVE = re.compile(r"^FAB\s*(\d{3,4})$", re.IGNORECASE)
RE_AERONAVE_BARE = re.compile(r"^(\d{3,4})$")
SECOES_VALIDAS = {"POR HORA", "POR POUSO", "POR CALENDÁRIO", "POR CALENDARIO", "POR DATA"}

REGISTRO = [
    {
        "operador": "CLA",
        "arquivo": OPERADORES_DIR / "CLA" / "Controle_de_Vencimentos_CLA_JUN_2026.xlsx",
        "tipo": "xlsx",
        "mes_fonte": "2026-06",
    },
    {
        "operador": "DACTA II",
        "arquivo": OPERADORES_DIR / "DACTA_II" / "Controle_de_Vencimentos_DACTAII_JUL_2026.ods",
        "tipo": "ods",
        "mes_fonte": "2026-07",
    },
    {
        # Confirmado pelo Wallace: a pasta "PAMA-LS" do Drive (maio) e a pasta
        # "BABR" (abril/junho) têm as mesmas aeronaves (2709, 2720, 2721) — é
        # a base BABR mesmo, só a pasta mudou de nome. 2704 é a única
        # aeronave que é de fato do PAMA-LS (ver registro seguinte).
        "operador": "BABR",
        "arquivo": OPERADORES_DIR / "BABR" / "Controle_de_Vencimentos_BABR_JUL_2026.xlsx",
        "tipo": "xlsx",
        "mes_fonte": "2026-07",
    },
    {
        "operador": "BABE",
        # A fonte de julho veio sem DISPONIBILIDADE nas duas linhas, portanto
        # não permite classificar o vencimento sem inventar valor. Pela regra
        # de busca, mantemos o último arquivo utilizável deste operador.
        "arquivo": OPERADORES_DIR / "BABE" / "Controle_de_Vencimentos_BABE_JUN_2026.xlsx",
        "tipo": "xlsx",
        "mes_fonte": "2026-06",
    },
    {
        # Formato padrão igual aos demais operadores; usamos o XLSX, não o
        # PDF duplicado da mesma pasta.
        "operador": "BACO",
        "arquivo": OPERADORES_DIR / "BACO" / "Controle_de_Vencimentos_BACO_JUL_2026.xlsx",
        "tipo": "xlsx",
        "mes_fonte": "2026-07",
    },
    {
        # Formato diferente dos demais: sem seções POR HORA/POUSO/CALENDÁRIO,
        # e a coluna AERONAVE fica vazia em cada linha porque a matrícula (só
        # a 2704) já vem fixada no cabeçalho/título do arquivo.
        "operador": "PAMA-LS",
        "arquivo": OPERADORES_DIR / "PAMA-LS-real" / "Vencimentos_2704_MAIO_2026.csv",
        "tipo": "csv_aeronave_fixa",
        "mes_fonte": "2026-05",
        "matricula_fixa": "2704",
        "linhas_cabecalho": 4,
    },
    {
        # Vem numa aba "VENCIMENTO" dentro do arquivo de Diagonal, com células
        # mescladas verticalmente e ordem de colunas própria.
        "operador": "BAMN",
        "arquivo": OPERADORES_DIR / "BAMN" / "Diagonal_de_Manutencao_C98_JULHO_2026.ods",
        "tipo": "ods_bamn",
        "mes_fonte": "2026-07",
        "aba": "VENCIMENTO",
    },
    {
        # Fonte íntegra de julho/2026. A ordem é a padrão, mas AERONAVE vem
        # sem prefixo "FAB" (só o número, como na BAMN).
        "operador": "BACG",
        "arquivo": OPERADORES_DIR / "BACG" / "Controle_de_Vencimentos_BACG_JUL_2026.xlsx",
        "tipo": "xlsx_aba_bare",
        "mes_fonte": "2026-07",
        "aba": "Controle de Vencimentos",
    },
    {
        # Arquivo único (Diagonal+MAPEM+Vencimento) com 3 abas — a aba
        # "Controle de Vencimento de Itens" usa a ordem de colunas padrão,
        # mas AERONAVE vem como número puro (int, sem "FAB"). Formatos de
        # valor extras encontrados aqui: horas com ponto de milhar
        # ("4.938:00"), horas anotadas com equivalente em meses entre
        # parênteses ("1.337:20 (19,8 M)"), "ANV S/ MOTOR"/"NÃO INSTALADO(A)"
        # (item não instalado, sem vencimento aplicável) e "VENCIDA" (vencido
        # sabido, mas sem valor informado). "JUL / 27" (mês/ano com espaços)
        # também apareceu na coluna de data.
        "operador": "BANT",
        "arquivo": OPERADORES_DIR / "BANT" / "DIAGONAL_E_VENC_ITENS_C98_JUL26_BANT.xlsx",
        "tipo": "xlsx_aba_bare",
        "mes_fonte": "2026-07",
        "aba": "Controle de Vencimento de Itens",
    },
]


def _ler_xlsx(caminho, aba=None):
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb[aba] if aba else wb[wb.sheetnames[0]]
    return [[ws.cell(row=r, column=c).value for c in range(1, 9)] for r in range(1, ws.max_row + 1)]


def _ler_ods(caminho):
    from odf.opendocument import load
    from odf.table import Table, TableCell, TableRow
    from odf.text import P

    def texto_celula(cell):
        return "".join(
            node.data for p in cell.getElementsByType(P) for node in p.childNodes if node.nodeType == node.TEXT_NODE
        )

    doc = load(str(caminho))
    tabela = doc.spreadsheet.getElementsByType(Table)[0]
    linhas = []
    for linha_odf in tabela.getElementsByType(TableRow):
        valores = []
        for cel in linha_odf.getElementsByType(TableCell):
            repete = int(cel.getAttribute("numbercolumnsrepeated") or 1)
            texto = texto_celula(cel) or None
            valores.extend([texto] * min(repete, 1))
        while len(valores) < 9:
            valores.append(None)
        linhas.append(valores[:9])
    return linhas


def _ler_ods_mesclado(caminho, nome_aba):
    """Lê uma aba de .ods "puxando pra baixo" o valor de células com
    mesclagem vertical (covered-table-cell) — necessário quando a
    nomenclatura de um item só é escrita na primeira linha do grupo e vale
    também pras linhas de baixo (ex.: BAMN)."""
    from odf.opendocument import load
    from odf.table import Table, TableRow
    from odf.text import P

    def texto_celula(cell):
        return "".join(
            node.data for p in cell.getElementsByType(P) for node in p.childNodes if node.nodeType == node.TEXT_NODE
        )

    doc = load(str(caminho))
    tabelas = [t for t in doc.spreadsheet.getElementsByType(Table) if t.getAttribute("name") == nome_aba]
    if not tabelas:
        return []
    tabela = tabelas[0]

    ultimo_valor_col = {}
    linhas = []
    for linha_odf in tabela.getElementsByType(TableRow):
        valores = []
        col_idx = 0
        for filho in linha_odf.childNodes:
            repete = int(filho.getAttribute("numbercolumnsrepeated") or 1)
            if filho.qname[1] == "covered-table-cell":
                for _ in range(repete):
                    valores.append(ultimo_valor_col.get(col_idx))
                    col_idx += 1
            else:
                texto = texto_celula(filho) or None
                for _ in range(min(repete, 1)):
                    ultimo_valor_col[col_idx] = texto
                    valores.append(texto)
                    col_idx += 1
        while len(valores) < 9:
            valores.append(None)
        linhas.append(valores[:9])
    return linhas


def _processar_linhas_bamn(linhas, operador, mes_fonte, arquivo_fonte, inconsistencias):
    """Ordem de colunas própria da BAMN: NOMENCLATURA/PN/SN/ESPECIALIDADE/
    AERONAVE/DISPONIBILIDADE/DATA/EMERG/OBS — e matrícula sem prefixo "FAB"."""
    itens = []
    secao_atual = None
    for linha in linhas:
        col0 = linha[0]
        if isinstance(col0, str) and col0.strip().upper() in SECOES_VALIDAS:
            secao_atual = col0.strip()
            continue

        nomenclatura, pn, sn, especialidade, aeronave = linha[0], linha[1], linha[2], linha[3], linha[4]
        disponibilidade, data_venc = linha[5], linha[6]

        if not isinstance(aeronave, str):
            continue
        m = RE_AERONAVE_BARE.match(aeronave.strip())
        if not m:
            continue
        matricula = m.group(1)

        tipo, valor_numerico, texto_original = classificar_disponibilidade(disponibilidade)
        if tipo in (None, "Desconhecido"):
            if disponibilidade not in (None, ""):
                inconsistencias.append(
                    f"{operador} ({arquivo_fonte}): DISPONIBILIDADE não reconhecida ({disponibilidade!r}) "
                    f"para FAB {matricula} / {nomenclatura} — item ignorado."
                )
            continue

        itens.append({
            "operador": operador,
            "mes_fonte": mes_fonte,
            "arquivo_fonte": arquivo_fonte,
            "secao_planilha": secao_atual,
            "especialidade": especialidade,
            "nomenclatura": nomenclatura,
            "pn": str(pn).strip() if pn is not None else None,
            "sn": str(sn).strip() if sn is not None else None,
            "matricula": matricula,
            "tipo_vencimento": tipo,
            "disponibilidade_valor": valor_numerico,
            "disponibilidade_texto": texto_original,
            "data_vencimento": parse_data_vencimento(data_venc),
            "vencido": vencido_de(tipo, valor_numerico),
        })
    return itens


def _ler_csv(caminho, linhas_cabecalho):
    import csv as csv_mod

    with open(caminho, encoding="utf-8") as f:
        linhas = list(csv_mod.reader(f))
    return linhas[linhas_cabecalho:]


def _processar_linhas_aeronave_fixa(linhas, operador, matricula_fixa, mes_fonte, arquivo_fonte, inconsistencias):
    itens = []
    for linha in linhas:
        if len(linha) < 9:
            continue
        especialidade, nomenclatura, pn, sn, _aeronave, disponibilidade, data_venc = linha[:7]
        if not nomenclatura.strip() or not disponibilidade.strip():
            continue

        tipo, valor_numerico, texto_original = classificar_disponibilidade(disponibilidade.strip())
        if tipo in (None, "Desconhecido"):
            inconsistencias.append(
                f"{operador} ({arquivo_fonte}): DISPONIBILIDADE não reconhecida ({disponibilidade!r}) "
                f"para FAB {matricula_fixa} / {nomenclatura} — item ignorado."
            )
            continue

        itens.append({
            "operador": operador,
            "mes_fonte": mes_fonte,
            "arquivo_fonte": arquivo_fonte,
            "secao_planilha": None,
            "especialidade": especialidade or None,
            "nomenclatura": nomenclatura,
            "pn": pn.strip() or None,
            "sn": sn.strip() or None,
            "matricula": matricula_fixa,
            "tipo_vencimento": tipo,
            "disponibilidade_valor": valor_numerico,
            "disponibilidade_texto": texto_original,
            "data_vencimento": parse_data_vencimento(data_venc.strip()) if data_venc.strip() else None,
            "vencido": vencido_de(tipo, valor_numerico),
        })
    return itens


def _processar_linhas(linhas, operador, mes_fonte, arquivo_fonte, inconsistencias, re_aeronave=RE_AERONAVE):
    itens = []
    secao_atual = None
    for linha in linhas:
        col0 = linha[0]
        if isinstance(col0, str) and col0.strip().upper() in SECOES_VALIDAS:
            secao_atual = col0.strip()
            continue

        especialidade, nomenclatura, pn, sn, aeronave = linha[0], linha[1], linha[2], linha[3], linha[4]
        disponibilidade, data_venc = linha[5], linha[6]

        if aeronave is None:
            continue
        # Aeronave pode vir como texto ("FAB2723") ou número puro (2703,
        # como na BANT) dependendo de como a fonte grava a célula.
        m = re_aeronave.match(str(aeronave).strip())
        if not m:
            continue
        matricula = m.group(1)

        tipo, valor_numerico, texto_original = classificar_disponibilidade(disponibilidade)
        if tipo in (None, "Desconhecido"):
            if disponibilidade not in (None, ""):
                inconsistencias.append(
                    f"{operador} ({arquivo_fonte}): DISPONIBILIDADE não reconhecida ({disponibilidade!r}) "
                    f"para FAB {matricula} / {nomenclatura} — item ignorado."
                )
            continue

        itens.append({
            "operador": operador,
            "mes_fonte": mes_fonte,
            "arquivo_fonte": arquivo_fonte,
            "secao_planilha": secao_atual,
            "especialidade": especialidade,
            "nomenclatura": nomenclatura,
            "pn": str(pn).strip() if pn is not None else None,
            "sn": str(sn).strip() if sn is not None else None,
            "matricula": matricula,
            "tipo_vencimento": tipo,
            "disponibilidade_valor": valor_numerico,
            "disponibilidade_texto": texto_original,
            "data_vencimento": parse_data_vencimento(data_venc),
            "vencido": vencido_de(tipo, valor_numerico),
        })
    return itens


def extrair():
    inconsistencias = []
    todos_itens = []

    for reg in REGISTRO:
        caminho = reg["arquivo"]
        if not caminho.exists():
            inconsistencias.append(f"{reg['operador']}: arquivo não encontrado em {caminho} — operador ignorado.")
            continue
        if reg["tipo"] == "xlsx":
            linhas = _ler_xlsx(caminho)
            itens = _processar_linhas(linhas, reg["operador"], reg["mes_fonte"], caminho.name, inconsistencias)
        elif reg["tipo"] == "ods":
            linhas = _ler_ods(caminho)
            itens = _processar_linhas(linhas, reg["operador"], reg["mes_fonte"], caminho.name, inconsistencias)
        elif reg["tipo"] == "csv_aeronave_fixa":
            linhas = _ler_csv(caminho, reg["linhas_cabecalho"])
            itens = _processar_linhas_aeronave_fixa(
                linhas, reg["operador"], reg["matricula_fixa"], reg["mes_fonte"], caminho.name, inconsistencias
            )
        elif reg["tipo"] == "csv_padrao_bare":
            linhas = _ler_csv(caminho, 1)
            itens = _processar_linhas(
                linhas, reg["operador"], reg["mes_fonte"], caminho.name, inconsistencias,
                re_aeronave=RE_AERONAVE_BARE,
            )
        elif reg["tipo"] == "xlsx_aba_bare":
            linhas = _ler_xlsx(caminho, aba=reg["aba"])
            itens = _processar_linhas(
                linhas, reg["operador"], reg["mes_fonte"], caminho.name, inconsistencias,
                re_aeronave=RE_AERONAVE_BARE,
            )
        else:  # ods_bamn
            linhas = _ler_ods_mesclado(caminho, reg["aba"])
            itens = _processar_linhas_bamn(linhas, reg["operador"], reg["mes_fonte"], caminho.name, inconsistencias)
        if not itens:
            inconsistencias.append(f"{reg['operador']}: nenhum item reconhecido em {caminho.name}.")
        todos_itens.extend(itens)

    df = pd.DataFrame(todos_itens)
    if not df.empty:
        # Células mescladas/repetidas em algumas fontes produzem cópias
        # integralmente idênticas. Mantê-las duplicaria itens e indicadores.
        df = df.drop_duplicates().reset_index(drop=True)
    return df, inconsistencias


def main():
    DADOS_TRATADOS.mkdir(parents=True, exist_ok=True)
    df, inconsistencias = extrair()

    destino = DADOS_TRATADOS / "base_vencimentos_operadores.xlsx"
    with pd.ExcelWriter(destino) as writer:
        df.to_excel(writer, index=False, sheet_name="Operadores")

    registrar_log(
        nome_execucao="extrair_vencimentos_operadores",
        arquivos_lidos=[str(r["arquivo"]) for r in REGISTRO],
        arquivos_gerados=[str(destino)],
        inconsistencias=inconsistencias,
    )

    print(f"{len(df)} itens de {df['operador'].nunique() if not df.empty else 0} operador(es) -> {destino}")
    if inconsistencias:
        print(f"{len(inconsistencias)} inconsistência(s) encontrada(s), ver log em 06_Logs/.")


if __name__ == "__main__":
    main()
