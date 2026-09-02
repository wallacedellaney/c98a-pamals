"""
Motores C-98 — extração da planilha "MOTORES C-98" (Google Sheets do Wallace,
pasta pessoal), 4 das 15 abas (escolhidas pelo Wallace em 2026-07-14 — as
demais são cenário/simulação, rascunho ou instrução, não dado):

- **SILOMS** ("Situacao" no arquivo tratado) — situação atual de cada motor
  (por OM/PN/SN), puxada do sistema SILOMS.
- **hélice** ("Helice") — mesma estrutura da SILOMS, só que de hélices.
- **Diagonal Nova** ("Diagonal") — projeção mês a mês (2025-2030) de quando
  cada motor vai bater TBO/HSI, com o comentário/nota da célula (quando
  existir) trazido junto — pedido do Wallace: "ja vi que tem comentarios
  dentro da caixas de tbo, hsi. vamos usar essas informacoes tb".

Gera 02_Dados_Tratados/base_motores_tratada.xlsx (3 abas: Situacao/Diagonal/
Helice). Ver 00_Instrucoes/motores.md — inclui as decisões de nomeação
das colunas ambíguas (3 colunas "DATA" repetidas com nomes diferentes na
planilha original de SILOMS/Helice — significado exato não confirmado com
o Wallace, nomeadas data_1/data_2 defensivamente).

2026-09-02: aba "OS" (ordens de serviço de motor em andamento) removida —
a fonte parou de ter essa aba (`Worksheet OS does not exist`, erro
recorrente desde 2026-08-26) e o Wallace confirmou que pode tirar de vez,
não só parar de dar erro (extração + aba "Ordens de Serviço" do site).

Fonte é uma planilha PESSOAL do Wallace (dono fred_o_m@hotmail.com),
compartilhada com a conta de serviço em 2026-07-15 — entrou na automação
de 2 em 2h (`atualizar_do_drive()`, ver shared/executar_atualizacao.py).
"""

from datetime import timedelta

import openpyxl
import pandas as pd

from common import BASES_ORIGINAIS, DADOS_TRATADOS, ESTADO_ATUALIZACOES, registrar_log
from shared import drive_sync, estado, horario
from shared.escrita_atomica import caminho_temporario

ARQUIVO_FONTE = BASES_ORIGINAIS / "Motores" / "MOTORES_C98.xlsx"

# Planilha pessoal do Wallace — compartilhada com a conta de serviço em
# 2026-07-15 ("ja compartilhei a planilha"), entrou na automação de 2 em 2h.
DRIVE_FILE_ID = "1UJDXA6jG4va51Tpnjd6DrV1kqPMnbY9w-TYlh8Ub0rM"

HISTORICO_SITUACAO = DADOS_TRATADOS / "historico_motores_situacao.csv"
COLUNAS_HISTORICO_SITUACAO = [
    "om", "pn", "sn", "matricula", "parcial_tso", "totais_tsn", "pct_tbo_voada",
    "tbo", "condicao", "motivo",
]

HISTORICO_DIAGONAL = DADOS_TRATADOS / "historico_motores_diagonal.csv"
COLUNAS_HISTORICO_DIAGONAL = ["serial", "anv", "ano", "mes", "evento", "comentario"]

# Índice de coluna (0-based) por campo — mapeado à mão a partir da estrutura
# real de cada aba (não são iguais entre si: SILOMS tem TBO antes de
# Matr.ANV, hélice tem a ordem invertida). "data_1"/"data_2" são as 2
# primeiras das 3 colunas "DATA" repetidas na fonte — significado exato não
# confirmado com o Wallace.
COL_SITUACAO = {
    "om": 6, "projeto": 7, "pn": 8, "tipo": 9, "fabricante": 10,
    "mnt_nivel_parque": 11, "estoque_utilizavel": 12, "estoque_reparavel": 13,
    "sn": 14, "controle": 15, "parcial_tso": 16, "totais_tsn": 17,
    "pct_tbo_voada": 18, "matricula": 19, "tbo": 20,
    "data_1": 21, "data_2": 22, "recolhimento": 23, "condicao": 24,
    "numero_doc": 25, "data_doc": 26, "motivo": 27,
}
LINHA_INICIO_SITUACAO = 4

COL_HELICE = {
    "om": 0, "projeto": 1, "pn": 2, "tipo": 3, "fabricante": 4,
    "mnt_nivel_parque": 5, "estoque_utilizavel": 6, "estoque_reparavel": 7,
    "sn": 8, "controle": 9, "parcial_tso": 10, "totais_tsn": 11,
    "pct_tbo_voada": 12, "tbo": 13, "matricula": 14,
    "data_1": 15, "data_2": 16, "recolhimento": 17, "condicao": 18,
    "numero_doc": 19, "data_doc": 20, "motivo": 21,
}
LINHA_INICIO_HELICE = 3

MESES_ABREV_FONTE = ["JAN", "FEV", "MAR", "ABR", "MAIO", "JUN", "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"]
COL_DIAGONAL_META = {"serial": 0, "anv": 1, "tso": 2, "hr_disp": 3, "voo_mensal": 4, "hr_fim_ano_anv": 5, "mes_disp": 6}
COL_DIAGONAL_PRIMEIRO_MES = 7
COL_DIAGONAL_ULTIMA_COLUNA = 79
LINHA_ANO_DIAGONAL = 2
LINHA_MES_DIAGONAL = 3
LINHA_INICIO_DIAGONAL = 4


def _numero(valor):
    """Colunas de horas de motor vêm como datetime.timedelta (Excel/Sheets
    trata "horas:minutos" como duração) — convertido pra float de horas.
    Já numérico passa direto; texto/vazio vira None."""
    if isinstance(valor, timedelta):
        return round(valor.total_seconds() / 3600, 2)
    if isinstance(valor, (int, float)):
        return float(valor)
    return None


def _texto(valor):
    if valor is None:
        return None
    if isinstance(valor, str):
        texto = valor.strip()
        return texto or None
    return valor


def _data(valor):
    return valor if hasattr(valor, "strftime") else None


def _texto_id(valor):
    """PN/SN/Matrícula às vezes vêm como número puro (ex.: 3044000.0) e às
    vezes como texto alfanumérico (ex.: "3104100-01") na mesma coluna — força
    string sempre, sem ".0" no final, pra não misturar tipo na mesma coluna
    (isso quebra o `st.dataframe`/Arrow e a ordenação, mesmo bug já visto em
    outras áreas — ver `ordenar_unicos` no Contrato 005)."""
    if valor is None:
        return None
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    texto = str(valor).strip()
    return texto or None


CAMPOS_HORA = {"parcial_tso", "totais_tsn", "tbo"}
CAMPOS_DATA_SITUACAO = {"data_1", "data_2", "data_doc"}
CAMPOS_ID = {"pn", "sn", "matricula", "numero_doc", "recolhimento", "solicitacao", "emergencia"}


def _extrair_situacao_generico(ws, colunas, linha_inicio):
    linhas = []
    for row in ws.iter_rows(min_row=linha_inicio, values_only=True):
        om = row[colunas["om"]] if colunas["om"] < len(row) else None
        if om in (None, ""):
            continue
        registro = {}
        for campo, idx in colunas.items():
            valor = row[idx] if idx < len(row) else None
            if campo in CAMPOS_HORA:
                registro[campo] = _numero(valor)
            elif campo in CAMPOS_DATA_SITUACAO:
                registro[campo] = _data(valor)
            elif campo in CAMPOS_ID:
                registro[campo] = _texto_id(valor)
            else:
                registro[campo] = _texto(valor)
        linhas.append(registro)
    return pd.DataFrame(linhas, columns=list(colunas.keys()))


def _numero_ou_none(valor):
    """'#N/A' (texto de fórmula do Sheets) e False (marcador de ANV vazio)
    viram None — o resto passa por `_numero` (converte timedelta em horas)."""
    if valor in (None, "", False) or (isinstance(valor, str) and valor.strip().upper() in ("#N/A", "#N/D")):
        return None
    return _numero(valor)


def _extrair_diagonal_metadados(wb):
    """1 linha por motor (serial) com os campos de planejamento já calculados
    na própria planilha: TSO, Hr disp (horas disponíveis até o próximo
    evento), Voo mensal (média mensal de horas de voo assumida) e Mês disp
    (= Hr disp / Voo mensal). Pedido do Wallace em 2026-07-14/15: expor
    "Voo mensal" como campo editável na Diagonal de Manutenção, pra poder
    simular "e se eu voar mais/menos por mês" sem mexer na planilha
    original (essa fica sempre fixa, seguindo o histórico real)."""
    ws = wb["Diagonal Nova"]
    linhas = []
    for row in ws.iter_rows(min_row=LINHA_INICIO_DIAGONAL, values_only=True):
        serial_valor = row[COL_DIAGONAL_META["serial"]] if COL_DIAGONAL_META["serial"] < len(row) else None
        if not isinstance(serial_valor, str):
            continue
        linhas.append({
            "serial": _texto(serial_valor),
            "anv": _numero_ou_none(row[COL_DIAGONAL_META["anv"]]),
            "tso": _numero_ou_none(row[COL_DIAGONAL_META["tso"]]),
            "hr_disp": _numero_ou_none(row[COL_DIAGONAL_META["hr_disp"]]),
            "voo_mensal": _numero_ou_none(row[COL_DIAGONAL_META["voo_mensal"]]),
            "hr_fim_ano_anv": _numero_ou_none(row[COL_DIAGONAL_META["hr_fim_ano_anv"]]),
            "mes_disp": _numero_ou_none(row[COL_DIAGONAL_META["mes_disp"]]),
        })
    return pd.DataFrame(linhas, columns=["serial", "anv", "tso", "hr_disp", "voo_mensal", "hr_fim_ano_anv", "mes_disp"])


def _extrair_diagonal(wb, inconsistencias):
    ws = wb["Diagonal Nova"]
    linha_ano = [c.value for c in ws[LINHA_ANO_DIAGONAL]]
    linha_mes = [c.value for c in ws[LINHA_MES_DIAGONAL]]

    mapa_mes = {}
    ano_atual = None
    for idx in range(COL_DIAGONAL_PRIMEIRO_MES, COL_DIAGONAL_ULTIMA_COLUNA):
        if idx < len(linha_ano) and linha_ano[idx] is not None:
            ano_atual = int(linha_ano[idx])
        mes_nome = str(linha_mes[idx]).strip() if idx < len(linha_mes) and linha_mes[idx] else None
        if mes_nome in MESES_ABREV_FONTE:
            mapa_mes[idx] = (ano_atual, MESES_ABREV_FONTE.index(mes_nome) + 1)

    linhas = []
    for row in ws.iter_rows(min_row=LINHA_INICIO_DIAGONAL):
        serial_valor = row[COL_DIAGONAL_META["serial"]].value if COL_DIAGONAL_META["serial"] < len(row) else None
        if not isinstance(serial_valor, str):
            continue
        anv_valor = row[COL_DIAGONAL_META["anv"]].value if COL_DIAGONAL_META["anv"] < len(row) else None
        for idx, (ano, mes) in mapa_mes.items():
            if idx >= len(row):
                continue
            cell = row[idx]
            if cell.value in (None, ""):
                continue
            linhas.append({
                "serial": _texto(serial_valor),
                "anv": anv_valor,
                "ano": ano,
                "mes": mes,
                # sempre string — a grade mistura marcadores de evento (TBO/
                # HSI/TBO*/X) com números e texto livre na mesma coluna;
                # sem isso a coluna fica com tipo misto e quebra a
                # serialização Arrow do st.dataframe.
                "evento": str(cell.value).strip() if cell.value is not None else None,
                "comentario": cell.comment.text.strip() if cell.comment else None,
            })
    return pd.DataFrame(linhas, columns=["serial", "anv", "ano", "mes", "evento", "comentario"])


COLUNAS_FINANCEIRO = ["pacote", "sigla_projeto", "atividade", "tarefa", "categoria", "valor_total"]

# Tabela financeira (Pacote/Sigla Projeto/Atividade/Tarefa/Categoria/Valor
# Total) da aba "Página7" da planilha de Motores — colunas M/R/S/T/U/AA,
# transcrita manualmente a partir de uma print do Wallace em 2026-07-28
# ("da coluna K até V tem valores financeiro").  Tarefas de motor por trás
# do "Para Motores" do MTA (Material HSI, Reparo de acessórios, Revisão
# Geral = TBO, Publicações etc.), com o mesmo "Pacote" (A/B/C/RAP) usado no
# MTA. **Fixa por decisão do Wallace em 2026-07-28** ("vai ser sempre
# aqueles dados, pode gravar, não vai mudar") — não é lida do Drive (nem
# por export de arquivo, que trunca essa aba por causa de colunas ocultas +
# filtro, nem pela API do Sheets, que devolveu a mesma tabela vazia — a aba
# provavelmente nunca chegou a ser compartilhada com a conta de serviço).
DADOS_FINANCEIRO_MOTORES = [
    ("A",   "C-98", "REQUISIÇÃO PUBLICAÇÃO",                    "Maintenance and Overhaul Collection PN 3077123 PT6A-34/114A", "MOTOR", 200000.00),
    ("A",   "C-98", "CNT 048/CELOG-PAMASP/2022 - PW - Material", "Material HSI PT6A-114A (1/2)",                               "MOTOR", 700000.00),
    ("B",   "C-98", "CNT 048/CELOG-PAMASP/2022 - PW - Material", "Material HSI PT6A-114A (2/2)",                               "MOTOR", 800000.00),
    ("A",   "C-98", "Requisição CABW",                           "Material reparo acessórios PT6A-114A (1/3)",                 "MOTOR", 400000.00),
    ("A",   "C-98", "Requisição CABW",                           "Material reparo acessórios PT6A-114A (2/3)",                 "MOTOR", 400000.00),
    ("A",   "C-98", "Requisição CABW",                           "Material reparo acessórios PT6A-114A (3/3)",                 "MOTOR", 400000.00),
    ("A",   "C-98", "CNT XXX/CELOG-PAMASP/2026 - Material",      "Material reparo PT6A-114A (1/3)",                            "MOTOR", 1000000.00),
    ("B",   "C-98", "CNT XXX/CELOG-PAMASP/2026 - Material",      "Material reparo PT6A-114A (2/3)",                            "MOTOR", 1000000.00),
    ("B",   "C-98", "CNT XXX/CELOG-PAMASP/2026 - Material",      "Material reparo PT6A-114A (3/3)",                            "MOTOR", 1000000.00),
    ("A",   "C-98", "REQUISIÇÃO PUBLICAÇÃO",                    "Publicações",                                                 "MOTOR", 563000.00),
    ("A",   "C-98", "REQUISIÇÃO PUBLICAÇÃO",                    "Publicações de Acessórios PT6A-114A",                         "MOTOR", 320000.00),
    ("A",   "C-98", "REQUISIÇÃO SERVIÇO",                       "Reparo de CTVR 2 EA (1/2)",                                   "MOTOR", 120000.00),
    ("C",   "C-98", "REQUISIÇÃO SERVIÇO",                       "Reparo de CTVR 2 EA (2/2)",                                   "MOTOR", 120000.00),
    ("A",   "C-98", "REQUISIÇÃO SERVIÇO",                       "Reparo de PTVR 2 EA (1/2)",                                   "MOTOR", 80000.00),
    ("C",   "C-98", "REQUISIÇÃO SERVIÇO",                       "Reparo de PTVR 2 EA (2/2)",                                   "MOTOR", 80000.00),
    ("RAP", "C-98", "CNT 031/CELOG-PAMASP/2022 - PW - RG",       "Revisão Geral PT6A-114A (1/6)",                              "MOTOR", 3800000.00),
    ("RAP", "C-98", "CNT 031/CELOG-PAMASP/2022 - PW - RG",       "Revisão Geral PT6A-114A (2/6)",                              "MOTOR", 3800000.00),
    ("A",   "C-98", "CNT 031/CELOG-PAMASP/2022 - PW - RG",       "Revisão Geral PT6A-114A (3/6)",                              "MOTOR", 3800000.00),
    ("A",   "C-98", "CNT XXX/CELOG-PAMASP/2026 - XXX - RG",      "Revisão Geral PT6A-114A (4/6)",                              "MOTOR", 3800000.00),
    ("B",   "C-98", "CNT XXX/CELOG-PAMASP/2026 - XXX - RG",      "Revisão Geral PT6A-114A (5/6)",                              "MOTOR", 3800000.00),
    ("B",   "C-98", "CNT XXX/CELOG-PAMASP/2026 - XXX - RG",      "Revisão Geral PT6A-114A (6/6)",                              "MOTOR", 3800000.00),
    ("A",   "C-98", "REQUISIÇÃO SERVIÇO",                       "RG FCU 3 EA (1/2)",                                           "MOTOR", 430000.00),
    ("A",   "C-98", "REQUISIÇÃO SERVIÇO",                       "RG FCU 3 EA (2/2)",                                           "MOTOR", 430000.00),
    ("A",   "C-98", "REQUISIÇÃO SERVIÇO",                       "RG Fuel Pump 5 EA (1/1)",                                     "MOTOR", 425000.00),
    ("A",   "C-98", "REQUISIÇÃO SERVIÇO",                       "RG IGNITION EXCITER 4 EA (1/1)",                             "MOTOR", 60000.00),
]


def _financeiro_existente():
    """Devolve a tabela financeira fixa (ver DADOS_FINANCEIRO_MOTORES) —
    não depende de arquivo nem de rede."""
    return pd.DataFrame(DADOS_FINANCEIRO_MOTORES, columns=COLUNAS_FINANCEIRO)


def extrair(df_financeiro=None):
    inconsistencias = []
    wb = openpyxl.load_workbook(ARQUIVO_FONTE, data_only=True)
    df_situacao = _extrair_situacao_generico(wb["SILOMS"], COL_SITUACAO, LINHA_INICIO_SITUACAO)
    df_helice = _extrair_situacao_generico(wb["hélice"], COL_HELICE, LINHA_INICIO_HELICE)
    df_diagonal = _extrair_diagonal(wb, inconsistencias)
    df_diagonal_meta = _extrair_diagonal_metadados(wb)
    if df_financeiro is None:
        df_financeiro = _financeiro_existente()
    return {
        "situacao": df_situacao, "diagonal": df_diagonal, "helice": df_helice,
        "diagonal_meta": df_diagonal_meta, "financeiro": df_financeiro,
    }, inconsistencias


def _registrar_historico_situacao(df_situacao):
    """Acrescenta o snapshot de hoje (1 linha por SN) — se já rodou hoje
    antes, substitui só as linhas de hoje (não duplica). Pedido do Wallace em
    2026-07-14: "vai ter historico pq vai ter atualizacao diaria" — grava
    toda vez que a extração roda (botão do site ou pedido na conversa),
    mesmo padrão de RAC/MTA/TPJL. Só existe história a partir do dia em que
    essa função passou a rodar."""
    hoje = horario.hoje_br().isoformat()
    novo = df_situacao[COLUNAS_HISTORICO_SITUACAO].copy()
    novo.insert(0, "data_snapshot", hoje)

    if HISTORICO_SITUACAO.exists():
        historico = pd.read_csv(HISTORICO_SITUACAO)
        historico = historico[historico["data_snapshot"] != hoje]
        historico = pd.concat([historico, novo], ignore_index=True)
    else:
        historico = novo
    historico.to_csv(HISTORICO_SITUACAO, index=False)


def _registrar_historico_diagonal(df_diagonal):
    """Acrescenta o snapshot de hoje dos eventos TBO/HSI/TBO* projetados (1
    linha por serial/ano/mês) — mesmo padrão da Situação. Pedido do Wallace
    em 2026-07-15: "mostrar na diagonal dos motores tb, um historico de
    evolucao" — a projeção pode mudar de um dia pro outro (mês empurrado,
    virou HSI em vez de TBO, comentário novo), então vale acompanhar."""
    hoje = horario.hoje_br().isoformat()
    eventos = df_diagonal[df_diagonal["evento"].isin({"TBO", "TBO*", "HSI"})]
    novo = eventos[COLUNAS_HISTORICO_DIAGONAL].copy()
    novo.insert(0, "data_snapshot", hoje)

    if HISTORICO_DIAGONAL.exists():
        historico = pd.read_csv(HISTORICO_DIAGONAL)
        historico = historico[historico["data_snapshot"] != hoje]
        historico = pd.concat([historico, novo], ignore_index=True)
    else:
        historico = novo
    historico.to_csv(HISTORICO_DIAGONAL, index=False)


def main(df_financeiro=None):
    DADOS_TRATADOS.mkdir(parents=True, exist_ok=True)
    dados, inconsistencias = extrair(df_financeiro=df_financeiro)

    destino = DADOS_TRATADOS / "base_motores_tratada.xlsx"
    with caminho_temporario(destino) as tmp:
        with pd.ExcelWriter(tmp) as writer:
            dados["situacao"].to_excel(writer, index=False, sheet_name="Situacao")
            dados["diagonal"].to_excel(writer, index=False, sheet_name="Diagonal")
            dados["helice"].to_excel(writer, index=False, sheet_name="Helice")
            dados["diagonal_meta"].to_excel(writer, index=False, sheet_name="DiagonalMeta")
            dados["financeiro"].to_excel(writer, index=False, sheet_name="Financeiro")

    _registrar_historico_situacao(dados["situacao"])
    _registrar_historico_diagonal(dados["diagonal"])

    registrar_log(
        nome_execucao="extrair_motores",
        arquivos_lidos=[str(ARQUIVO_FONTE)],
        arquivos_gerados=[str(destino)],
        inconsistencias=inconsistencias,
    )

    for chave, df in dados.items():
        print(f"{chave}: {len(df)} linha(s) -> {destino} (aba {chave.capitalize()})")
    if inconsistencias:
        print(f"{len(inconsistencias)} inconsistência(s) encontrada(s), ver log em 06_Logs/.")

    return dados


def atualizar_do_drive():
    """Busca a versão mais recente direto do Google Drive, sobrescreve a
    cópia local e reprocessa. Ver 00_Instrucoes/atualizacoes.md.

    A tabela financeira (aba "Página7") não entra nessa busca — é fixa
    (ver DADOS_FINANCEIRO_MOTORES), `main()` já usa ela por padrão."""
    try:
        metadados = drive_sync.obter_metadados(DRIVE_FILE_ID)
        conteudo = drive_sync.baixar_arquivo(DRIVE_FILE_ID, exportar_como=drive_sync.XLSX_MIME)
        ARQUIVO_FONTE.parent.mkdir(parents=True, exist_ok=True)
        ARQUIVO_FONTE.write_bytes(conteudo)

        dados = main()
        estado.atualizar_estado(
            ESTADO_ATUALIZACOES, "motores",
            remote_modified_time=metadados["modifiedTime"],
            local_updated_at=horario.agora_br().isoformat(),
            status="atualizado",
            record_count=len(dados["situacao"]),
            last_error=None,
        )
    except Exception as e:
        estado.atualizar_estado(ESTADO_ATUALIZACOES, "motores", status="erro", last_error=str(e))
        raise
    return estado.obter_entrada(ESTADO_ATUALIZACOES, "motores")


if __name__ == "__main__":
    import sys
    if "--atualizar-do-drive" in sys.argv:
        atualizar_do_drive()
    else:
        main()
