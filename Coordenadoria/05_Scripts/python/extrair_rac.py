"""
Carrega a matriz aeronave x material da planilha "Análise crítica de emergências
C-98 2026" (aba 'Rac') e gera 02_Dados_Tratados/base_rac_tratada.xlsx.

Estrutura da aba (ver 00_Instrucoes/rac.md):
* Linha 1: matrícula da aeronave, por coluna (a partir da coluna G). A cor de
  fundo real vem de formatação CONDICIONAL do Google Sheets baseada na linha 3
  (contagem de pendências daquela aeronave) — não é uma cor estática na célula:
  - Fundo estático coral (#E06666)     -> Sem condições (fora do contrato)
  - Sem fundo estático, linha 3 = 0    -> Montada
  - Sem fundo estático, linha 3 > 0    -> Desmontada
  - Fonte vermelha (#FF0000) -> fora do contrato (independente da disponibilidade)
  - Fonte padrão (preta)     -> dentro do contrato
* Linha 2: unidade/esquadrão da aeronave.
* Linha 3: contagem de pendências por aeronave (usada só para decidir
  Montada/Desmontada — recalculamos o nosso próprio total a partir da matriz).
* Linha 5 em diante: uma linha por PN (Núm, PN, Nomenclatura, Sum, Cont nas
  colunas A-E); da coluna G em diante, o valor é a quantidade que falta
  daquele PN para aquela aeronave (vazio/zero = não falta).
"""

from datetime import datetime

import openpyxl
import pandas as pd

from common import BASES_ORIGINAIS, DADOS_TRATADOS, ESTADO_ATUALIZACOES, registrar_log
from shared import drive_sync, estado

FONTE = BASES_ORIGINAIS / "RAC" / "Analise critica de emergencias C-98 2026 (Google Sheets).xlsx"
ABA = "Rac"

# Ver 00_Instrucoes/atualizacoes.md — planilha Google nativa "Análise crítica
# de emergências C-98 2026", aba "Rac".
DRIVE_FILE_ID = "1o8supQLcHkC1WZZCZDAtuRKGB_VUlQ8qBlYj7racsGQ"

# Um snapshot por dia (item a item) das pendências, acumulado desde
# 2026-07-06 — não existe histórico anterior a essa data, pois o RAC sempre
# sobrescreveu a cópia local até então. Ver 00_Instrucoes/rac.md.
HISTORICO_RAC = DADOS_TRATADOS / "historico_rac.csv"

PRIMEIRA_COLUNA_AERONAVE = 7  # coluna G
LINHA_CONTAGEM_PENDENCIAS = 3
PRIMEIRA_LINHA_PN = 5

FUNDO_INDISPONIVEL = "FFE06666"
FONTE_FORA_CONTRATO = "FFFF0000"


def _cor(color):
    if color is None:
        return None
    if color.type == "rgb":
        return color.rgb
    return None  # cor por tema (ex.: preto padrão) — tratada como "dentro do contrato"


def extrair():
    inconsistencias = []
    wb = openpyxl.load_workbook(FONTE, data_only=True)
    ws = wb[ABA]

    # --- Aeronaves (cabeçalho, colunas G em diante) ---
    aeronaves = []
    col = PRIMEIRA_COLUNA_AERONAVE
    while ws.cell(row=1, column=col).value is not None:
        cel_matricula = ws.cell(row=1, column=col)
        matricula = str(cel_matricula.value)
        unidade = ws.cell(row=2, column=col).value

        fundo = _cor(cel_matricula.fill.fgColor) if cel_matricula.fill else None
        if fundo == FUNDO_INDISPONIVEL:
            disponibilidade = "Sem condições"
        else:
            contagem_pendencias = ws.cell(row=LINHA_CONTAGEM_PENDENCIAS, column=col).value or 0
            disponibilidade = "Montada" if contagem_pendencias == 0 else "Desmontada"

        fonte = _cor(cel_matricula.font.color) if cel_matricula.font and cel_matricula.font.color else None
        # Fundo vermelho/coral (indisponível) é sempre fora do contrato, mesmo
        # com fonte preta; fonte vermelha também marca fora do contrato mesmo
        # com fundo verde (ver rac.md).
        if fonte == FONTE_FORA_CONTRATO or fundo == FUNDO_INDISPONIVEL:
            contrato = "Fora do contrato"
        else:
            contrato = "Dentro do contrato"

        aeronaves.append({
            "coluna": col,
            "matricula": matricula,
            "unidade": unidade,
            "disponibilidade": disponibilidade,
            "contrato": contrato,
        })
        col += 1

    # --- Pendências (uma linha por PN x aeronave com falta > 0) ---
    pendencias = []
    linha = PRIMEIRA_LINHA_PN
    while True:
        pn = ws.cell(row=linha, column=2).value
        if pn is None:
            # tolera até 3 linhas em branco seguidas antes de parar (planilha pode ter espaços)
            if all(ws.cell(row=r, column=2).value is None for r in range(linha, linha + 3)):
                break
            linha += 1
            continue

        nomenclatura = ws.cell(row=linha, column=3).value
        for aeronave in aeronaves:
            valor = ws.cell(row=linha, column=aeronave["coluna"]).value
            if valor:
                pendencias.append({
                    "matricula": aeronave["matricula"],
                    "unidade": aeronave["unidade"],
                    "pn": pn,
                    "nomenclatura": nomenclatura,
                    "quantidade_faltante": valor,
                })
        linha += 1

    df_aeronaves = pd.DataFrame(aeronaves).drop(columns=["coluna"])
    df_pendencias = pd.DataFrame(pendencias)

    resumo = df_pendencias.groupby("matricula").agg(
        total_pendencias=("pn", "count"),
        soma_unidades_faltantes=("quantidade_faltante", "sum"),
    ).reset_index()
    df_aeronaves = df_aeronaves.merge(resumo, on="matricula", how="left")
    df_aeronaves["total_pendencias"] = df_aeronaves["total_pendencias"].fillna(0).astype(int)
    df_aeronaves["soma_unidades_faltantes"] = df_aeronaves["soma_unidades_faltantes"].fillna(0).astype(int)

    return df_aeronaves, df_pendencias, inconsistencias


def main():
    DADOS_TRATADOS.mkdir(parents=True, exist_ok=True)
    df_aeronaves, df_pendencias, inconsistencias = extrair()

    destino = DADOS_TRATADOS / "base_rac_tratada.xlsx"
    with pd.ExcelWriter(destino) as writer:
        df_aeronaves.to_excel(writer, index=False, sheet_name="Aeronaves")
        df_pendencias.to_excel(writer, index=False, sheet_name="Pendencias")

    registrar_log(
        nome_execucao="extrair_rac",
        arquivos_lidos=[str(FONTE)],
        arquivos_gerados=[str(destino)],
        inconsistencias=inconsistencias,
    )

    print(f"{len(df_aeronaves)} aeronaves, {len(df_pendencias)} pendências -> {destino}")
    if inconsistencias:
        print(f"{len(inconsistencias)} inconsistência(s) encontrada(s), ver log em 06_Logs/.")

    return df_aeronaves, df_pendencias


def _registrar_historico(df_pendencias):
    """Acrescenta o snapshot de hoje ao histórico item a item — se já rodou
    hoje antes, substitui só as linhas de hoje (não duplica)."""
    hoje = datetime.now().date().isoformat()
    novo = df_pendencias[["matricula", "unidade", "pn", "nomenclatura", "quantidade_faltante"]].copy()
    novo.insert(0, "data", hoje)

    if HISTORICO_RAC.exists():
        historico = pd.read_csv(HISTORICO_RAC, dtype={"matricula": str})
        historico = historico[historico["data"] != hoje]
        historico = pd.concat([historico, novo], ignore_index=True)
    else:
        historico = novo
    historico.to_csv(HISTORICO_RAC, index=False)


def atualizar_do_drive():
    """Busca a versão mais recente direto do Google Drive, sobrescreve a
    cópia local e reprocessa. Ver 00_Instrucoes/atualizacoes.md."""
    try:
        metadados = drive_sync.obter_metadados(DRIVE_FILE_ID)
        conteudo = drive_sync.baixar_arquivo(DRIVE_FILE_ID, exportar_como=drive_sync.XLSX_MIME)
        FONTE.parent.mkdir(parents=True, exist_ok=True)
        FONTE.write_bytes(conteudo)
        df_aeronaves, df_pendencias = main()
        _registrar_historico(df_pendencias)
        estado.atualizar_estado(
            ESTADO_ATUALIZACOES, "rac",
            remote_modified_time=metadados["modifiedTime"],
            local_updated_at=datetime.now().isoformat(),
            status="atualizado",
            record_count=len(df_aeronaves),
            last_error=None,
        )
    except Exception as e:
        estado.atualizar_estado(ESTADO_ATUALIZACOES, "rac", status="erro", last_error=str(e))
        raise
    return estado.obter_entrada(ESTADO_ATUALIZACOES, "rac")


if __name__ == "__main__":
    import sys
    if "--atualizar-do-drive" in sys.argv:
        atualizar_do_drive()
    else:
        main()
