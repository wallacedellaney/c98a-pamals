"""Leitura compartilhada das grades de "Diagonal de Manutenção" (projeção de
inspeções/indisponibilidade por aeronave ao longo do tempo) — ver
00_Instrucoes/diagonal_manutencao.md.

Cada operador desenha a grade de um jeito (semanal com "Semana N", semanal com
faixa de dias "1-11", mensal com 2 linhas por aeronave), mas o sinal comum é
sempre o mesmo: quando a linha "INSP. PROG." / "IMPACTO DIAG." (ou equivalente)
tem texto numa coluna de tempo, aquela aeronave fica indisponível naquele
período por causa do que está escrito ali (confirmado pelo Wallace). Este
módulo generaliza essa leitura em vez de codar um parser por operador.
"""

import re
import warnings
from pathlib import Path

import openpyxl

MESES = {
    "JANEIRO": 1, "FEVEREIRO": 2, "MARÇO": 3, "MARCO": 3, "ABRIL": 4, "MAIO": 5, "JUNHO": 6,
    "JULHO": 7, "AGOSTO": 8, "SETEMBRO": 9, "OUTUBRO": 10, "NOVEMBRO": 11, "DEZEMBRO": 12,
}
RE_FAB = re.compile(r"^FAB\s*(\d{3,4})$", re.IGNORECASE)
RE_BARE = re.compile(r"^(\d{3,4})$")
RE_ANO = re.compile(r"\b(20\d{2})\b")
RE_INFO = re.compile(r"INSP\.?\s*PROG|IMPACTO\s*DIAG", re.IGNORECASE)

# Textos de modelo/exemplo que aparecem em planilhas não preenchidas (ex.: CLA
# manda o arquivo praticamente em branco, só com "XX:XX"/"INSP-XXX" de
# exemplo) — não são eventos reais, não podem virar indisponibilidade.
RE_PLACEHOLDER = re.compile(r"^(INSP-[A-Z]{2,4}|XX:XX|X{2,})$", re.IGNORECASE)

# "IS" (Indisponível por Suprimento) e "sem motor" são status/condição da
# aeronave, não inspeção programada (confirmado pelo Wallace: "só inspeção
# mesmo") — mesmo aparecendo na linha "INSP. PROG." de algumas fontes, não
# entram como evento de manutenção programada.
RE_NAO_PROGRAMADO = re.compile(r"^(AERONAVE\s*)?(IS\s*(\([^)]*\))?|SEM\s*MOTOR)$", re.IGNORECASE)


def eh_programado(texto):
    """True se o texto representa uma inspeção programada de verdade (não um
    status como "IS"/"sem motor"). Para textos combinados de várias partes
    (BAMN junta com "; "), verifica cada parte."""
    partes = [p.strip() for p in texto.split(";") if p.strip()]
    partes_programadas = [p for p in partes if not RE_NAO_PROGRAMADO.match(p)]
    return "; ".join(partes_programadas) if partes_programadas else None


def matricula_de(valor):
    if valor is None:
        return None
    texto = str(valor).strip()
    m = RE_FAB.match(texto)
    if m:
        return m.group(1)
    m = RE_BARE.match(texto)
    if m:
        return m.group(1)
    return None


def _ler_matriz_xlsx(caminho, sheet):
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb[sheet]
    max_col = ws.max_column
    max_row = ws.max_row
    return [[ws.cell(row=r, column=c).value for c in range(1, max_col + 1)] for r in range(1, max_row + 1)]


def _ler_matriz_ods(caminho, sheet):
    from odf.opendocument import load
    from odf.table import Table, TableRow, TableCell
    from odf.text import P

    def texto_celula(cell):
        return "".join(
            node.data for p in cell.getElementsByType(P) for node in p.childNodes if node.nodeType == node.TEXT_NODE
        )

    doc = load(str(caminho))
    tabelas = [t for t in doc.spreadsheet.getElementsByType(Table) if t.getAttribute("name").strip() == sheet]
    if not tabelas:
        return []
    tabela = tabelas[0]
    linhas = []
    max_col = 0
    for linha_odf in tabela.getElementsByType(TableRow):
        valores = []
        for cel in linha_odf.getElementsByType(TableCell):
            repete = int(cel.getAttribute("numbercolumnsrepeated") or 1)
            texto = texto_celula(cel) or None
            valores.extend([texto] * min(repete, 1))
        linhas.append(valores)
        max_col = max(max_col, len(valores))
    return [linha + [None] * (max_col - len(linha)) for linha in linhas]


def ler_grade_generica(caminho, sheet):
    """Lê uma grade estilo BANT/DACTA II/BABR/BABE/BACO/CLA: cabeçalho com
    nomes de mês (repetido uma vez no topo ou por bloco de aeronave), sub-
    cabeçalho opcional "Semana N", e blocos de 2 linhas por aeronave (uma
    linha "HORAS VOO" e uma "INSP. PROG."/"IMPACTO DIAG." — a matrícula pode
    estar em qualquer uma das duas linhas do bloco, varia por operador).

    Retorna lista de (matricula, ano_mes 'YYYY-MM', semana_label ou None, motivo).
    """
    caminho = Path(caminho)
    if caminho.suffix.lower() == ".ods":
        matriz = _ler_matriz_ods(caminho, sheet)
    else:
        matriz = _ler_matriz_xlsx(caminho, sheet)

    max_row = len(matriz)
    max_col = max((len(linha) for linha in matriz), default=0)

    periodo_por_col = {}
    ano_atual = None
    eventos = []

    r = 1
    while r <= max_row:
        linha_atual = matriz[r - 1] if r - 1 < len(matriz) else []
        valores = linha_atual + [None] * (max_col - len(linha_atual))

        for v in valores:
            if isinstance(v, str):
                m = RE_ANO.search(v)
                if m:
                    ano_atual = m.group(1)

        meses_encontrados = {
            c: MESES[str(v).strip().upper()]
            for c, v in enumerate(valores, start=1)
            if isinstance(v, str) and str(v).strip().upper() in MESES
        }
        if meses_encontrados:
            linha_prox = matriz[r] if r < len(matriz) else []
            prox = linha_prox + [None] * (max_col - len(linha_prox))
            tem_semana = any(isinstance(v, str) and "SEMANA" in str(v).upper() for v in prox)
            cols_mes = sorted(meses_encontrados.keys())
            for i, col_ini in enumerate(cols_mes):
                col_fim = cols_mes[i + 1] - 1 if i + 1 < len(cols_mes) else max_col
                mes_num = meses_encontrados[col_ini]
                for col in range(col_ini, col_fim + 1):
                    semana_label = None
                    if tem_semana and isinstance(prox[col - 1], str) and "SEMANA" in str(prox[col - 1]).upper():
                        semana_label = str(prox[col - 1]).strip()
                    periodo_por_col[col] = (ano_atual, mes_num, semana_label)
            r += 2 if tem_semana else 1
            continue

        info_idx = None
        for c, v in enumerate(valores, start=1):
            if isinstance(v, str) and RE_INFO.search(v):
                info_idx = c
                break
        if info_idx is not None:
            matricula = None
            for rr in (r, r - 1, r + 1):
                if rr < 1 or rr > max_row:
                    continue
                linha_rr = matriz[rr - 1] if rr - 1 < len(matriz) else []
                for c in range(1, min(info_idx + 1, 6)):
                    valor_c = linha_rr[c - 1] if c - 1 < len(linha_rr) else None
                    matricula = matricula_de(valor_c)
                    if matricula:
                        break
                if matricula:
                    break
            if matricula:
                for col in range(info_idx + 1, max_col + 1):
                    val = valores[col - 1] if col - 1 < len(valores) else None
                    if val in (None, ""):
                        continue
                    texto = str(val).strip()
                    if RE_PLACEHOLDER.match(texto):
                        continue
                    texto_programado = eh_programado(texto)
                    if texto_programado is None:
                        continue
                    ano, mes, semana = periodo_por_col.get(col, (None, None, None))
                    if ano is None or mes is None:
                        continue
                    eventos.append((matricula, f"{ano}-{mes:02d}", semana, texto_programado))
        r += 1
    return eventos


def periodo_para_datas(ano_mes, semana_label):
    """Converte 'YYYY-MM' + rótulo de semana (ou None) num (data_inicio, data_fim)
    aproximado, só pra plotar no Gantt — não é uma data exata de calendário
    confirmada pela fonte, é uma aproximação de qual semana do mês."""
    from datetime import date
    import calendar

    ano, mes = int(ano_mes[:4]), int(ano_mes[5:7])
    ultimo_dia = calendar.monthrange(ano, mes)[1]
    if not semana_label:
        return date(ano, mes, 1), date(ano, mes, ultimo_dia)

    m = re.search(r"(\d+)", semana_label)
    if not m:
        return date(ano, mes, 1), date(ano, mes, ultimo_dia)
    n = int(m.group(1))
    inicio = min((n - 1) * 7 + 1, ultimo_dia)
    fim = min(n * 7, ultimo_dia) if n < 4 else ultimo_dia
    return date(ano, mes, inicio), date(ano, mes, fim)
