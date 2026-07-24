"""
Extrai a planilha de pagamentos (aba 'CONTROLE DE PAGAMENTOS', complementada pela
aba 'EMPENHOS') e gera 02_Dados_Tratados/base_pagamentos_tratada.xlsx.

Regras de tratamento vêm de 00_Instrucoes/pagamentos.md.
"""

import re
import warnings

import openpyxl
import pandas as pd

from common import XLSX_PAGAMENTOS, DADOS_TRATADOS, ESTADO_ATUALIZACOES, registrar_log
from shared import drive_sync, estado, horario

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# Ver 00_Instrucoes/atualizacoes.md — planilha Google nativa
# "005/CELOG-PAMALS/2025 online", dono wallacedellaney@gmail.com.
DRIVE_FILE_ID = "1zV_SQKlcXVYeaqCbV0X-PiWnzdzOZPt5esaXp_4k6_o"

# Snapshot diário dos lançamentos — mesmo padrão de Emergências/RAC/MTA/TPJL.
# Sem um ID único na planilha; a chave usada pra comparar dia a dia é a
# combinação (tipo_registro, modulo, referencia, numero_recibo,
# numero_nota_fiscal). Só existe histórico a partir de quando essa gravação
# começou. Ver 00_Instrucoes/analise_periodo.md.
HISTORICO_PAGAMENTOS = DADOS_TRATADOS / "historico_pagamentos.csv"
COLUNAS_HISTORICO = [
    "tipo_registro", "modulo", "referencia", "numero_recibo", "numero_nota_fiscal",
    "valor_nfs", "faturado", "pendente", "situacao",
]

# Colunas C..N da aba Controle de Pagamentos (mesma estrutura nos dois blocos de
# lançamentos: mensal e por módulo/orçamento). Ver pagamentos.md.
COLS_PAGAMENTO = [
    "modulo", "referencia", "numero_recibo", "numero_nota_fiscal", "data",
    "empenho", "observacao", "vencimento", "valor_nfs", "ordem_pagamento",
    "faturado", "pendente",
]


def extrair_contrato(ws):
    return {
        "numero_contrato": ws["D5"].value,
        "ug_executora": ws["E5"].value,
        "ug_responsavel": ws["F5"].value,
        "ug_fiscalizadora": ws["G5"].value,
        "status": ws["J5"].value,
        "valor_total_contrato": ws["N5"].value,
        "valor_a_empenhar": ws["O5"].value,
        "valor_empenhado": ws["P5"].value,
        "valor_liquidado": ws["Q5"].value,
        "saldo_a_faturar": ws["R5"].value,
        "fornecedor": ws["S5"].value,
    }


def parse_valor_monetario(valor):
    """Alguns lançamentos vêm como número, outros como texto 'R$ 1.234,56' (ver pagamentos.md)."""
    if pd.isna(valor):
        return None, False
    if isinstance(valor, (int, float)):
        return float(valor), False
    texto = str(valor).replace("R$", "").strip().replace(".", "").replace(",", ".")
    try:
        return float(texto), True
    except ValueError:
        return None, True


def padronizar_valores_monetarios(df, inconsistencias):
    for coluna in ("valor_nfs", "faturado", "pendente"):
        valores_convertidos = []
        for idx, valor in df[coluna].items():
            novo_valor, era_texto = parse_valor_monetario(valor)
            valores_convertidos.append(novo_valor)
            if era_texto:
                inconsistencias.append(
                    f"Coluna '{coluna}' na referência {df.at[idx, 'referencia']} veio como texto "
                    f"('{valor}') em vez de número; convertido para {novo_valor}."
                )
        df[coluna] = valores_convertidos
    return df


REGEX_ORDEM_PAGAMENTO = re.compile(r"^\d{4}NP\d+$")


def situacao_pagamento(row):
    # Nem todo valor não vazio em "Ordem de Pagamento" é uma ordem de
    # verdade — às vezes vem só o texto "Faturado" (ainda sem número de
    # ordem), o que não é "Pago" de fato. Bug real visto em 2026-07-18: o
    # lançamento "Módulo III – Orçamento 19 GPS" (NF 1882) tinha
    # ordem_pagamento="Faturado" e caía como "Pago" por engano — só um
    # valor no formato real de ordem (ex.: "2026NP401058") conta como Pago.
    ordem = row["ordem_pagamento"]
    if pd.notna(ordem) and REGEX_ORDEM_PAGAMENTO.match(str(ordem).strip()):
        return "Pago"
    if pd.notna(row["valor_nfs"]) or pd.notna(row["faturado"]):
        return "Faturado, aguardando pagamento"
    if pd.notna(row["pendente"]):
        return "Pendente"
    return "Sem lançamento"


def localizar_headers(ws):
    """Acha as linhas de cabeçalho ('Módulo | Referência | ...') dos 2 blocos de
    lançamentos (mensal, depois por módulo/orçamento) procurando o texto
    'Referência' na coluna D, em vez de assumir números de linha fixos.

    Corrige um bug real visto em 2026-07-18: o lançamento "Módulo III –
    Orçamento 19 GPS" (NF 1882, empenho 2025NE001065) ficou de fora dos
    Pagamentos porque o bloco "por módulo/orçamento" tinha fim fixo em
    ultima_linha=39 no código — quando esa linha nova foi acrescentada na
    planilha (linha 40), a extração simplesmente não a via. Como o Wallace
    vai continuar acrescentando linhas com o tempo, descobrir os limites de
    cada bloco dinamicamente evita ter que atualizar um número fixo no
    código toda vez."""
    return [linha for linha in range(1, ws.max_row + 1) if ws.cell(row=linha, column=4).value == "Referência"]


def fim_do_bloco(ws, primeira_linha):
    """A partir de `primeira_linha`, acha a última linha com algum dado nas
    colunas C..N (o bloco "por módulo/orçamento" é o último da planilha —
    termina numa sequência de linhas totalmente em branco)."""
    ultima_com_dado = primeira_linha - 1
    linha = primeira_linha
    while linha <= ws.max_row:
        valores = [ws.cell(row=linha, column=c).value for c in range(3, 15)]
        if all(v is None for v in valores):
            break
        ultima_com_dado = linha
        linha += 1
    return ultima_com_dado


def extrair_bloco(ws, primeira_linha, ultima_linha):
    linhas = []
    for row in ws.iter_rows(min_row=primeira_linha, max_row=ultima_linha, min_col=3, max_col=14, values_only=True):
        linhas.append(dict(zip(COLS_PAGAMENTO, row)))
    df = pd.DataFrame(linhas)
    df["modulo"] = df["modulo"].ffill()
    return df


def extrair_empenhos(wb):
    ws = wb["EMPENHOS"]
    empenhos = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        ne = row[0]
        if not ne:
            continue
        empenhos[str(ne).strip()] = {
            "valor_empenhado": row[10],   # coluna K
            "saldo": row[11],             # coluna L
            "responsavel": row[26],       # coluna AA
            "justificativa": row[27],     # coluna AB
        }
    return empenhos


def empenhos_para_dataframe(empenhos):
    linhas = [{"numero_empenho": ne, **dados} for ne, dados in empenhos.items()]
    return pd.DataFrame(linhas, columns=["numero_empenho", "valor_empenhado", "saldo", "responsavel", "justificativa"])


def cruzar_empenhos(df, empenhos):
    saldos, valores, responsaveis, justificativas = [], [], [], []
    inconsistencias = []

    for _, row in df.iterrows():
        empenho_cell = row["empenho"]
        nes = [n.strip() for n in re.split(r"[,\n]", str(empenho_cell)) if n.strip()] if pd.notna(empenho_cell) else []

        encontrados = [empenhos[ne] for ne in nes if ne in empenhos]
        nao_encontrados = [ne for ne in nes if ne and ne not in empenhos]
        for ne in nao_encontrados:
            inconsistencias.append(f"Empenho '{ne}' (referência {row['referencia']}) não encontrado na aba EMPENHOS.")

        saldos.append(sum(e["saldo"] for e in encontrados if e["saldo"]) or None)
        valores.append(sum(e["valor_empenhado"] for e in encontrados if e["valor_empenhado"]) or None)
        responsaveis.append("; ".join(sorted({e["responsavel"] for e in encontrados if e["responsavel"]})) or None)
        justificativas.append("; ".join(sorted({e["justificativa"] for e in encontrados if e["justificativa"]})) or None)

    df["empenho_saldo"] = saldos
    df["empenho_valor_empenhado"] = valores
    df["empenho_responsavel"] = responsaveis
    df["empenho_justificativa"] = justificativas
    return df, inconsistencias


def main():
    DADOS_TRATADOS.mkdir(parents=True, exist_ok=True)
    inconsistencias = []

    wb = openpyxl.load_workbook(XLSX_PAGAMENTOS, data_only=True)
    ws = wb["CONTROLE DE PAGAMENTOS"]

    contrato = extrair_contrato(ws)

    headers = localizar_headers(ws)
    if len(headers) != 2:
        raise ValueError(
            f"Esperava 2 blocos de lançamentos (cabeçalho 'Referência' na coluna D), "
            f"encontrei {len(headers)} nas linhas {headers}. Conferir a planilha original."
        )
    header_mensal, header_modulo = headers

    bloco_mensal = extrair_bloco(ws, header_mensal + 1, header_modulo - 1)
    bloco_mensal["tipo_registro"] = "mensal"

    bloco_modulo = extrair_bloco(ws, header_modulo + 1, fim_do_bloco(ws, header_modulo + 1))
    bloco_modulo["tipo_registro"] = "orcamento"

    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", category=FutureWarning)
        df = pd.concat([bloco_mensal, bloco_modulo], ignore_index=True)
    df = df[df[["numero_recibo", "numero_nota_fiscal", "valor_nfs", "faturado", "pendente"]].notna().any(axis=1)]
    df = df.reset_index(drop=True)

    df = padronizar_valores_monetarios(df, inconsistencias)
    df["situacao"] = df.apply(situacao_pagamento, axis=1)

    empenhos = extrair_empenhos(wb)
    df, inconsistencias_empenho = cruzar_empenhos(df, empenhos)
    inconsistencias.extend(inconsistencias_empenho)
    df_empenhos = empenhos_para_dataframe(empenhos)

    destino = DADOS_TRATADOS / "base_pagamentos_tratada.xlsx"
    with pd.ExcelWriter(destino) as writer:
        pd.DataFrame([contrato]).to_excel(writer, index=False, sheet_name="Contrato")
        df.to_excel(writer, index=False, sheet_name="Pagamentos")
        df_empenhos.to_excel(writer, index=False, sheet_name="Empenhos")

    registrar_log(
        nome_execucao="extrair_pagamentos",
        arquivos_lidos=[str(XLSX_PAGAMENTOS)],
        arquivos_gerados=[str(destino)],
        inconsistencias=inconsistencias,
        proximas_acoes=["Revisar empenhos não encontrados na aba EMPENHOS."] if inconsistencias else None,
    )

    print(f"{len(df)} lançamentos de pagamento extraídos -> {destino}")
    if inconsistencias:
        print(f"{len(inconsistencias)} inconsistência(s) encontrada(s), ver log em 06_Logs/.")

    return df


def _registrar_historico(df):
    """Acrescenta o snapshot de hoje (1 linha por lançamento) — se já
    rodou hoje antes, substitui só as linhas de hoje (não duplica)."""
    hoje = horario.hoje_br().isoformat()
    novo = df[COLUNAS_HISTORICO].copy()
    novo.insert(0, "data_snapshot", hoje)

    if HISTORICO_PAGAMENTOS.exists():
        historico = pd.read_csv(HISTORICO_PAGAMENTOS)
        historico = historico[historico["data_snapshot"] != hoje]
        historico = pd.concat([historico, novo], ignore_index=True)
    else:
        historico = novo
    historico.to_csv(HISTORICO_PAGAMENTOS, index=False)


def atualizar_do_drive():
    """Busca a versão mais recente direto do Google Drive, sobrescreve a
    cópia local e reprocessa. Ver 00_Instrucoes/atualizacoes.md."""
    try:
        metadados = drive_sync.obter_metadados(DRIVE_FILE_ID)
        conteudo = drive_sync.baixar_arquivo(DRIVE_FILE_ID, exportar_como=drive_sync.XLSX_MIME)
        XLSX_PAGAMENTOS.parent.mkdir(parents=True, exist_ok=True)
        XLSX_PAGAMENTOS.write_bytes(conteudo)
        df = main()
        _registrar_historico(df)
        estado.atualizar_estado(
            ESTADO_ATUALIZACOES, "pagamentos",
            remote_modified_time=metadados["modifiedTime"],
            local_updated_at=horario.agora_br().isoformat(),
            status="atualizado",
            record_count=len(df),
            last_error=None,
        )
    except Exception as e:
        estado.atualizar_estado(ESTADO_ATUALIZACOES, "pagamentos", status="erro", last_error=str(e))
        raise
    return estado.obter_entrada(ESTADO_ATUALIZACOES, "pagamentos")


if __name__ == "__main__":
    import sys
    if "--atualizar-do-drive" in sys.argv:
        atualizar_do_drive()
    else:
        main()
