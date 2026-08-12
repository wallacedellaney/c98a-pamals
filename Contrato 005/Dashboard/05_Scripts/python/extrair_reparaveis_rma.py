"""Complementa o "Controle de Reparáveis" (planilha geral) com dados da RMA
"em andamento" do mês, pra OS que a empresa (VEE ONE) já devolveu de
verdade mas cuja burocracia ainda não fechou na planilha geral — pedido do
Wallace, 2026-08-12: "coloquei hoje [a RMA em andamento de julho]... tem
uns OS que foram entregues (a burocracia provavelmente ainda tá aberta)...
busca essas OS, [...] atualiza onde que tá, data de devolução pela empresa,
número do recibo, só deixar registrado que essa informação veio da RMA de
julho da empresa". Ampliado no mesmo dia: "pode puxar da aba 1.10 todas as
OS tb, [...] atualiza o status de TODAS OS da nossa planilha [...] saber
onde que tá, saber o recibo, saber quando eles entregaram e a fonte" — não
fica mais restrito às OS que a aba 1.8 lista como devolvidas NO MÊS; puxa
a aba 1.10 inteira (todo o controle acumulado de OS, não só do mês).

Fonte: mesmo arquivo "RMA em andamento {MÊS}.xlsx" / "Pré RMA C-98
{Mês}-26.xlsx" já usado pelo Financeiro (RMA) e pela Apresentação (RMA) —
ver `gerar_ata_reuniao._localizar_pasta_mes`/`_baixar_rma_em_andamento`
(mesma pasta do Drive "Fechamentos mensais" > ano > mês).

Fonte dos 3 campos: aba **1.10** "Controle de Ordens de Serviço abertas até
o mês de referência" — pra cada OS (coluna "Nº da OS"), tem "Data da
devolução", "N° do Recibo" e "Operador" (pra onde foi entregue), quando
preenchidos (a maioria das OS ainda em aberto de verdade fica em branco
nos 3 — normal, não é inconsistência). Toda OS da 1.10 com pelo menos um
dos 3 campos preenchido vira uma linha complementada, com a fonte ("RMA
{Mês}/{Ano}") — nunca sobrescreve a planilha geral (só complementa, ver
`02_Dados_Tratados/reparaveis_complemento_rma.xlsx`, mesclado por cima na
exibição do dashboard, não na base tratada).

Acumula por mês (upsert por OS+mes_referencia) — rodar de novo no mesmo mês
substitui as linhas desse mês; rodar num mês novo soma, sem apagar o
histórico de meses anteriores (e se uma OS aparecer complementada em mais
de um mês, o dashboard usa a mais recente — ver `_mesclar_complemento_rma`
em `reparaveis.py`).

**Condenados (aba 1.9, 2026-08-12)** — Wallace: "na aba 1.9 tem os
devolvidos condenados, busca nas OS aberto eles e acrescenta essa
informação". A aba **1.9** "Materiais reparáveis condenados" tem, pra cada
OS ("Nº da OS atrelada de Abertura"), a data de devolução do SN original e
o "Motivo de Condenação" — informação que a 1.10 não tem (uma OS condenada
pode não aparecer lá com data/recibo preenchidos). Toda OS da 1.9 vira
`condenado=True` + `motivo_condenacao`, mesclada na mesma linha da OS (se
ela já tinha dado da 1.10) ou como linha nova (se só existe na 1.9).
"""

import io
from datetime import date

import openpyxl
import pandas as pd

from common import DADOS_TRATADOS, registrar_log
from shared import horario

DESTINO = DADOS_TRATADOS / "reparaveis_complemento_rma.xlsx"

MESES_PT = [
    "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
]

# "PAMALS" (sem hífen) é como a RMA escreve o operador PAMA-LS — mesma
# variação de grafia já documentada em `contrato005/secoes/reparaveis.py`
# (LOCAIS_ENTREGUES) pra planilha geral, que usa "PAMA-LS" com hífen.
# Normaliza aqui pra bater com o mesmo valor usado lá (senão o item não
# entraria em LOCAIS_ENTREGUES por diferença pontual de grafia).
NORMALIZACAO_OPERADOR = {"PAMALS": "PAMA-LS", "PAMA LS": "PAMA-LS"}

# Números de OS reais têm 10 dígitos (ex.: 3040249831) — limiar bem abaixo
# disso só pra descartar leitura errada de linha de rodapé/total (ex.: "Total
# de itens devolvidos no período: 4 itens." — o "4" cai por acidente na
# mesma coluna da OS quando lido célula a célula).
OS_MINIMO_VALIDO = 100_000


def _valor_texto(v):
    if v is None:
        return None
    texto = str(v).strip()
    return texto or None


def _parse_data_condenacao(v):
    """Aceita datetime real ou texto "DD/MM/AAAA" — inclusive com barra
    duplicada por erro de digitação na origem (ex.: "22//07/2026", visto na
    aba 1.9 de julho/2026). Não inventa: formato não reconhecido vira None."""
    if hasattr(v, "date"):
        return v.date()
    if not isinstance(v, str):
        return None
    partes = [p for p in v.strip().split("/") if p]
    if len(partes) != 3:
        return None
    try:
        d, m, y = (int(p) for p in partes)
        return date(y, m, d)
    except ValueError:
        return None


def _extrair_condenados(wb, nome_arquivo):
    """Lê a aba 1.9 ("Materiais reparáveis condenados"). Devolve dict
    {numero_os: {"data_devolucao": date|None, "motivo": str|None}}."""
    if "1.9" not in wb.sheetnames:
        return {}
    ws = wb["1.9"]
    condenados = {}
    for r in range(6, ws.max_row + 1):
        nos = ws.cell(row=r, column=8).value  # "Nº da OS atrelada de Abertura"
        if not isinstance(nos, (int, float)) or nos < OS_MINIMO_VALIDO:
            continue
        condenados[int(nos)] = {
            "data_devolucao": _parse_data_condenacao(ws.cell(row=r, column=9).value),
            "motivo": _valor_texto(ws.cell(row=r, column=11).value),
        }
    return condenados


def extrair(conteudo_bytes, ano, mes, nome_arquivo):
    """Lê a aba 1.10 inteira (todo o controle de OS, não só as do mês) +
    1.9 (condenados) do conteúdo (bytes) da RMA em andamento. Devolve (df,
    inconsistencias) — df com colunas: os, mes_referencia, ano_referencia,
    data_devolucao_empresa, data_devolucao_empresa_texto, onde_se_encontra,
    recibo, fonte, entregue_no_mes, condenado, motivo_condenacao,
    arquivo_fonte. Toda OS com pelo menos um dos 3 campos
    (devolução/recibo/operador) preenchido na 1.10, OU que aparece na 1.9
    (condenada), vira uma linha — OS ainda sem nenhum dos 3 e fora da 1.9 é
    a maioria (ainda em aberto de verdade) e é ignorada silenciosamente,
    não é inconsistência.

    `entregue_no_mes` (bool) — pedido do Wallace, 2026-08-12: "oq foi
    entregue no mês são da 1.8, aí vc deixa claro lá tb" — distingue a OS
    que a aba **1.8** "Materiais reparáveis devolvidos no mês de
    referência" confirma como devolvida NESTE mês (True) de uma OS que só
    tem o dado histórico na 1.10, de um mês anterior (False) — a `fonte`
    reflete isso no texto ("... (entregue no mês)" x "... (histórico)")."""
    wb = openpyxl.load_workbook(io.BytesIO(conteudo_bytes), data_only=True)
    inconsistencias = []

    if "1.10" not in wb.sheetnames:
        inconsistencias.append(f"{nome_arquivo}: aba 1.10 não encontrada — nada extraído.")
        return pd.DataFrame(), inconsistencias

    os_entregues_no_mes = set()
    if "1.8" in wb.sheetnames:
        ws88 = wb["1.8"]
        for r in range(6, ws88.max_row + 1):
            numero_os = ws88.cell(row=r, column=5).value
            if isinstance(numero_os, (int, float)) and numero_os >= OS_MINIMO_VALIDO:
                os_entregues_no_mes.add(int(numero_os))
    else:
        inconsistencias.append(f"{nome_arquivo}: aba 1.8 não encontrada — não dá pra marcar quais OS foram entregues neste mês específico.")

    ws810 = wb["1.10"]
    mes_ano = f"{MESES_PT[mes - 1]}/{ano}"
    registros = {}
    for r in range(6, ws810.max_row + 1):
        nos = ws810.cell(row=r, column=7).value
        if not isinstance(nos, (int, float)) or nos < OS_MINIMO_VALIDO:
            continue
        numero_os = int(nos)

        data_devolucao_raw = ws810.cell(row=r, column=10).value
        onde = _valor_texto(ws810.cell(row=r, column=13).value)
        if onde and onde.upper() in NORMALIZACAO_OPERADOR:
            onde = NORMALIZACAO_OPERADOR[onde.upper()]
        recibo = _valor_texto(ws810.cell(row=r, column=12).value)

        data_devolucao = data_devolucao_raw.date() if hasattr(data_devolucao_raw, "date") else None
        data_devolucao_texto = None if data_devolucao else _valor_texto(data_devolucao_raw)

        if data_devolucao is None and data_devolucao_texto is None and onde is None and recibo is None:
            continue  # OS ainda em aberto de verdade, nem na própria RMA tem info — normal, maioria dos casos

        entregue_no_mes = numero_os in os_entregues_no_mes
        registros[numero_os] = {
            "os": str(numero_os),
            "mes_referencia": mes,
            "ano_referencia": ano,
            "data_devolucao_empresa": data_devolucao,
            "data_devolucao_empresa_texto": data_devolucao_texto,
            "onde_se_encontra": onde,
            "recibo": recibo,
            "entregue_no_mes": entregue_no_mes,
            "condenado": False,
            "motivo_condenacao": None,
            "arquivo_fonte": nome_arquivo,
        }

    # Aba 1.9 — "Materiais reparáveis condenados" (pedido do Wallace,
    # 2026-08-12: "na aba 1.9 tem os devolvidos condenados, busca nas OS
    # aberto eles e acrescenta essa informação"). Mescla na mesma linha da
    # OS (se já veio da 1.10) ou cria uma linha nova (só existia na 1.9).
    condenados = _extrair_condenados(wb, nome_arquivo)
    for numero_os, info in condenados.items():
        registro = registros.get(numero_os)
        if registro is None:
            registro = {
                "os": str(numero_os),
                "mes_referencia": mes,
                "ano_referencia": ano,
                "data_devolucao_empresa": None,
                "data_devolucao_empresa_texto": None,
                "onde_se_encontra": None,
                "recibo": None,
                "entregue_no_mes": numero_os in os_entregues_no_mes,
                "condenado": False,
                "motivo_condenacao": None,
                "arquivo_fonte": nome_arquivo,
            }
            registros[numero_os] = registro
        registro["condenado"] = True
        registro["motivo_condenacao"] = info["motivo"]
        if registro["data_devolucao_empresa"] is None and info["data_devolucao"] is not None:
            registro["data_devolucao_empresa"] = info["data_devolucao"]

    linhas = []
    for registro in registros.values():
        fonte = f"RMA {mes_ano} (entregue no mês)" if registro["entregue_no_mes"] else f"RMA {mes_ano} (histórico)"
        if registro["condenado"]:
            fonte += " — condenado"
        registro["fonte"] = fonte
        linhas.append(registro)

    return pd.DataFrame(linhas), inconsistencias


def atualizar_do_mes(ano, mes):
    import gerar_ata_reuniao as gar
    from shared import drive_sync

    drive_sync.garantir_credencial_arquivo()
    arquivos_pasta = gar._localizar_pasta_mes(ano, mes)
    conteudo, nome_arquivo = gar._baixar_rma_em_andamento(arquivos_pasta)

    df_novo, inconsistencias = extrair(conteudo, ano, mes, nome_arquivo)

    if DESTINO.exists():
        anterior = pd.read_excel(DESTINO, dtype={"os": str})
        anterior = anterior[
            ~((anterior["mes_referencia"] == mes) & (anterior["ano_referencia"] == ano))
        ]
        completo = pd.concat([anterior, df_novo], ignore_index=True) if not df_novo.empty else anterior
    else:
        completo = df_novo

    DADOS_TRATADOS.mkdir(parents=True, exist_ok=True)
    completo.to_excel(DESTINO, index=False)

    registrar_log(
        nome_execucao="extrair_reparaveis_rma",
        arquivos_lidos=[nome_arquivo],
        arquivos_gerados=[str(DESTINO)],
        inconsistencias=inconsistencias,
    )

    return {
        "arquivo": nome_arquivo,
        "os_complementadas_no_mes": len(df_novo),
        "os_entregues_no_mes": int(df_novo["entregue_no_mes"].sum()) if not df_novo.empty else 0,
        "os_historico": int((~df_novo["entregue_no_mes"]).sum()) if not df_novo.empty else 0,
        "os_condenadas": int(df_novo["condenado"].sum()) if not df_novo.empty else 0,
        "total_acumulado": len(completo),
        "inconsistencias": len(inconsistencias),
        "atualizado_em": horario.hoje_br().isoformat(),
    }


def main():
    hoje = horario.hoje_br()
    resultado = atualizar_do_mes(hoje.year, hoje.month)
    print(resultado)


if __name__ == "__main__":
    main()
