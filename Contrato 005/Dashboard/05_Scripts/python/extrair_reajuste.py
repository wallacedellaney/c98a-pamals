"""
Carrega a "Planilha Demonstrativa -2_Reajuste - Contrato 005_CELOG-PAMALS_2025"
(arquivo .xlsx real enviado ao Drive pelo Wallace, não Planilha Google
nativa) e gera 02_Dados_Tratados/base_reajuste_tratada.xlsx — aba
"Reajuste" no site. Ver 00_Instrucoes/reajuste.md.

A planilha tem 4 abas: "Valor do Contrato" e "Cronogroma Físico Financeiro"
(as 2 usadas aqui) + "Cópia de Valor do Contrato"/"Cópia de Cronogroma
Físico Fina" (cópias antigas guardadas pelo próprio Wallace, ignoradas).

Estrutura de "Valor do Contrato" (confirmada célula a célula em
2026-07-16): uma sequência de indicadores escalares (rótulo na coluna A,
valor na coluna C — alguns rótulos repetem, ex. "Módulo 1" aparece em
3 momentos diferentes: valor de assinatura, executado até 08/10/25, e
executado 08/10/25-08/10/26) intercalada com 2 tabelas de Notas Fiscais
(rótulo "Evento" no cabeçalho) — uma antes do 1° Reajuste, outra depois.
Extraído por BUSCA DE RÓTULO (texto exato na coluna A), não por número de
linha fixo — mais resistente a o Wallace inserir/remover linha na
planilha pessoal dele.
"""

from datetime import datetime

import openpyxl
import pandas as pd

from common import BASES_ORIGINAIS, DADOS_TRATADOS, ESTADO_ATUALIZACOES, registrar_log
from shared import drive_sync, estado

ARQUIVO_FONTE = (
    BASES_ORIGINAIS / "Reajuste"
    / "Planilha Demonstrativa -2_Reajuste - Contrato 005_CELOG-PAMALS_2025.xlsx"
)
DRIVE_FILE_ID = "1R32r4rscXTYGe98R1AFUUG8hqZD-GaWY"

ABA_VALOR = "Valor do Contrato"
# Nome real da aba na planilha do Wallace tem um erro de digitação
# ("Cronogroma", não "Cronograma") — mantido como está pra bater com a
# fonte de verdade.
ABA_CRONOGRAMA = "Cronogroma Físico Financeiro"

COLUNAS_NF = [
    "evento", "nota_fiscal", "valor_total_nf", "descricao_nf", "modulo",
    "orcamento", "autorizacao", "aprovacao", "emissao_nf", "vencimento_nf",
]


def _texto(v):
    if v is None:
        return None
    t = str(v).strip()
    return t if t and t != "-" else None


def _numero(v):
    if v is None or isinstance(v, str):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _valor_monetario(v):
    """Quase todo valor já vem numérico (openpyxl aplica a fórmula/formato).
    Uma célula real da fonte (Nota Fiscal de Módulo 3, Ofício 13/1218) veio
    como texto "R$ 58.868,35" em vez de número — só essa, tratada aqui pra
    não virar coluna com tipo misto (float + str), que quebra o Arrow no
    Streamlit."""
    if v is None:
        return None
    if isinstance(v, str):
        t = v.strip().replace("R$", "").strip()
        if not t or t == "-":
            return None
        t = t.replace(".", "").replace(",", ".")
        try:
            return float(t)
        except ValueError:
            return None
    return float(v)


def _linhas_coluna_a(ws, max_row):
    return [(r, ws.cell(row=r, column=1).value) for r in range(1, max_row + 1)]


def _extrair_indicadores(ws, max_row):
    """Todos os rótulos escalares da coluna A com valor na coluna C, na
    ordem em que aparecem na planilha — preserva rótulos repetidos (ex.
    "Módulo 1" em 3 seções diferentes) em vez de tentar adivinhar um nome
    único por rótulo, pra não inventar significado que a planilha não deixa
    explícito sozinha."""
    linhas = []
    # "Nota Fiscal" fica de fora daqui — já vira uma linha própria e mais
    # completa (com descrição/módulo/datas) na tabela _extrair_notas_fiscais.
    ignorar = {"Evento", "Descrição", "Mês", "Nota Fiscal"}
    for r, label in _linhas_coluna_a(ws, max_row):
        label = _texto(label)
        if not label or label in ignorar:
            continue
        valor = ws.cell(row=r, column=3).value
        if valor is None:
            continue
        if isinstance(valor, str) and valor.strip() in ("", "-"):
            continue
        linhas.append({"linha": r, "indicador": label, "valor": valor})
    return pd.DataFrame(linhas)


def _extrair_notas_fiscais(ws, max_row):
    """As 2 tabelas de Nota Fiscal (linha de cabeçalho = coluna A "Evento")
    — a 1ª ocorrência é sempre antes do 1° Reajuste, a 2ª depois (ordem
    cronológica de cima pra baixo na planilha)."""
    cabecalhos = [r for r, label in _linhas_coluna_a(ws, max_row) if _texto(label) == "Evento"]
    linhas = []
    for i, r_cabecalho in enumerate(cabecalhos):
        periodo = "Antes do 1° Reajuste" if i == 0 else "Após o 1° Reajuste"
        r = r_cabecalho + 1
        while r <= max_row:
            evento = _texto(ws.cell(row=r, column=1).value)
            valor_nf = ws.cell(row=r, column=3).value
            if evento != "Nota Fiscal" or valor_nf is None:
                break
            linhas.append({
                "periodo": periodo,
                "evento": evento,
                "nota_fiscal": _texto(ws.cell(row=r, column=2).value),
                "valor_total_nf": _valor_monetario(valor_nf),
                "descricao_nf": _texto(ws.cell(row=r, column=4).value),
                "modulo": _texto(ws.cell(row=r, column=5).value),
                "orcamento": _texto(ws.cell(row=r, column=6).value),
                "autorizacao": _texto(ws.cell(row=r, column=7).value),
                "aprovacao": ws.cell(row=r, column=8).value,
                "emissao_nf": ws.cell(row=r, column=9).value,
                "vencimento_nf": ws.cell(row=r, column=10).value,
            })
            r += 1
    return pd.DataFrame(linhas)


def _extrair_cronograma_mensal(ws):
    """Linhas mês x módulo (colunas A-D) até o marcador "(1) Valores
    Execução Autorizada" — pula a linha "1° Reajuste" (só um separador
    visual, sem dado de mês)."""
    linhas = []
    reajuste_aplicado = False
    for r in range(3, ws.max_row + 1):
        col_a = ws.cell(row=r, column=1).value
        if isinstance(col_a, str) and "Valores Execução Autorizada" in col_a:
            break
        if isinstance(col_a, str) and "Reajuste" in col_a and not isinstance(col_a, datetime):
            reajuste_aplicado = True
            continue
        if not isinstance(col_a, datetime):
            continue
        linhas.append({
            "mes": col_a,
            "modulo_1": _numero(ws.cell(row=r, column=2).value),
            "modulo_2": _numero(ws.cell(row=r, column=3).value),
            "modulo_3": _numero(ws.cell(row=r, column=4).value),
            "apos_1_reajuste": reajuste_aplicado,
        })
    return pd.DataFrame(linhas)


def _extrair_cronograma_resumo(ws):
    """Tabela final "Descrição / Módulo 1 / Módulo 2 / Módulo 3" (Proposta
    Comercial, Total Executado, Saldo sem Reajuste, Saldo/Valor por Módulo
    após 1° Reajuste, Valor total a executar/do contrato após 1°
    reajuste)."""
    r_cabecalho = None
    for r, label in _linhas_coluna_a(ws, ws.max_row):
        if _texto(label) == "Descrição" and _texto(ws.cell(row=r, column=2).value) == "Módulo 1":
            r_cabecalho = r
            break
    if r_cabecalho is None:
        return pd.DataFrame()

    linhas = []
    for r in range(r_cabecalho + 1, ws.max_row + 1):
        descricao = _texto(ws.cell(row=r, column=1).value)
        if not descricao:
            break
        linhas.append({
            "descricao": descricao,
            "modulo_1": _numero(ws.cell(row=r, column=2).value),
            "modulo_2": _numero(ws.cell(row=r, column=3).value),
            "modulo_3": _numero(ws.cell(row=r, column=4).value),
        })
    return pd.DataFrame(linhas)


def extrair():
    wb = openpyxl.load_workbook(ARQUIVO_FONTE, data_only=True)
    ws_valor = wb[ABA_VALOR]
    ws_cron = wb[ABA_CRONOGRAMA]

    indicadores = _extrair_indicadores(ws_valor, ws_valor.max_row)
    notas_fiscais = _extrair_notas_fiscais(ws_valor, ws_valor.max_row)
    cronograma_mensal = _extrair_cronograma_mensal(ws_cron)
    cronograma_resumo = _extrair_cronograma_resumo(ws_cron)

    return {
        "indicadores": indicadores,
        "notas_fiscais": notas_fiscais,
        "cronograma_mensal": cronograma_mensal,
        "cronograma_resumo": cronograma_resumo,
    }


def main():
    DADOS_TRATADOS.mkdir(parents=True, exist_ok=True)
    dados = extrair()

    destino = DADOS_TRATADOS / "base_reajuste_tratada.xlsx"
    with pd.ExcelWriter(destino) as writer:
        dados["indicadores"].to_excel(writer, index=False, sheet_name="Indicadores")
        dados["notas_fiscais"].to_excel(writer, index=False, sheet_name="NotasFiscais")
        dados["cronograma_mensal"].to_excel(writer, index=False, sheet_name="CronogramaMensal")
        dados["cronograma_resumo"].to_excel(writer, index=False, sheet_name="CronogramaResumo")

    registrar_log(
        nome_execucao="extrair_reajuste",
        arquivos_lidos=[str(ARQUIVO_FONTE)],
        arquivos_gerados=[str(destino)],
        inconsistencias=[],
    )

    print(f"{len(dados['indicadores'])} indicador(es), {len(dados['notas_fiscais'])} nota(s) fiscal(is), "
          f"{len(dados['cronograma_mensal'])} mês(es) de cronograma -> {destino}")
    return dados


def atualizar_do_drive():
    """Busca a versão mais recente direto do Google Drive, sobrescreve a
    cópia local e reprocessa. Ver 00_Instrucoes/atualizacoes.md."""
    try:
        metadados = drive_sync.obter_metadados(DRIVE_FILE_ID)
        conteudo = drive_sync.baixar_arquivo(DRIVE_FILE_ID)
        ARQUIVO_FONTE.parent.mkdir(parents=True, exist_ok=True)
        ARQUIVO_FONTE.write_bytes(conteudo)
        dados = main()
        estado.atualizar_estado(
            ESTADO_ATUALIZACOES, "reajuste",
            remote_modified_time=metadados["modifiedTime"],
            local_updated_at=datetime.now().isoformat(),
            status="atualizado",
            record_count=len(dados["indicadores"]),
            last_error=None,
        )
    except Exception as e:
        estado.atualizar_estado(ESTADO_ATUALIZACOES, "reajuste", status="erro", last_error=str(e))
        raise
    return estado.obter_entrada(ESTADO_ATUALIZACOES, "reajuste")


if __name__ == "__main__":
    import sys
    if "--atualizar-do-drive" in sys.argv:
        atualizar_do_drive()
    else:
        main()
